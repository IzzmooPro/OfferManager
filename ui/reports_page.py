"""Raporlama sayfası — analiz tabloları ve grafikler."""
import logging
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFrame, QComboBox, QTableWidget, QTableWidgetItem,
    QHeaderView, QFileDialog, QMessageBox,
)
from PySide6.QtCore import Qt, QEvent
from services.report_service import ReportService
from ui.widgets._row_hover_delegate import RowHoverDelegate
from core.constants import SYM_MAP
from core.formatting import fmt_money, fmt_number
from core.date_utils import to_display_date

logger = logging.getLogger("reports_page")


class ReportsPage(QWidget):
    def __init__(self):
        super().__init__()
        self._svc = ReportService()
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        hdr = QFrame()
        hdr.setObjectName("toolbar")
        hdr.setFixedHeight(64)
        hl = QHBoxLayout(hdr)
        hl.setContentsMargins(16, 12, 16, 12)
        title = QLabel("Raporlar")
        title.setStyleSheet("font-size:14pt;font-weight:700;")
        hl.addWidget(title)
        hl.addStretch()

        self._report_combo = QComboBox()
        self._report_combo.setMinimumHeight(34)
        self._report_combo.setMinimumWidth(220)
        self._report_combo.addItems([
            "Aylık Ciro Raporu",
            "Müşteri Sıralaması",
            "Ürün Sıralaması",
            "Teklif Dönüşüm Oranı",
        ])
        self._report_combo.currentIndexChanged.connect(self._generate)
        hl.addWidget(self._report_combo)

        hl.addSpacing(8)
        export_btn = QPushButton("Excel'e Aktar")
        export_btn.setObjectName("secondary")
        export_btn.setFixedHeight(34)
        export_btn.clicked.connect(self._export)
        hl.addWidget(export_btn)
        root.addWidget(hdr)

        body = QVBoxLayout()
        body.setContentsMargins(16, 12, 16, 12)
        body.setSpacing(8)

        self._summary_label = QLabel("")
        self._summary_label.setObjectName("hint_label")
        self._summary_label.setWordWrap(True)
        body.addWidget(self._summary_label)

        self._table = QTableWidget()
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.verticalHeader().setVisible(False)
        self._table.setAlternatingRowColors(False)
        # Diğer sayfalarla aynı tam satır hover davranışı
        self._hover_delegate = RowHoverDelegate(self)
        self._table.setItemDelegate(self._hover_delegate)
        self._table.setMouseTracking(True)
        self._table.viewport().setMouseTracking(True)
        self._table.viewport().installEventFilter(self)
        body.addWidget(self._table, 1)
        root.addLayout(body, 1)

    def eventFilter(self, obj, event):
        if obj is self._table.viewport():
            if event.type() == QEvent.Type.MouseMove:
                idx = self._table.indexAt(event.position().toPoint())
                row = idx.row() if idx.isValid() else -1
                if row != self._hover_delegate._hovered_row:
                    self._hover_delegate.set_hovered_row(row)
                    self._table.viewport().update()
            elif event.type() == QEvent.Type.Leave:
                self._hover_delegate.set_hovered_row(-1)
                self._table.viewport().update()
        return super().eventFilter(obj, event)

    def on_enter(self):
        self._generate()

    def _generate(self):
        idx = self._report_combo.currentIndex()
        self._reset_table()
        try:
            if idx == 0:
                self._report_monthly_revenue()
            elif idx == 1:
                self._report_customer_ranking()
            elif idx == 2:
                self._report_product_ranking()
            elif idx == 3:
                self._report_conversion()
        except Exception as e:
            logger.error("Rapor oluşturma hatası: %s", e, exc_info=True)
            self._summary_label.setText(f"Rapor oluşturulamadı: {e}")

    def _reset_table(self):
        self._table.clear()
        self._table.setRowCount(0)
        self._table.setColumnCount(0)

    def _setup_columns(self, headers: list):
        """Tüm sütunları Stretch yap — ekranı doldursun."""
        self._table.setColumnCount(len(headers))
        self._table.setHorizontalHeaderLabels(headers)
        hh = self._table.horizontalHeader()
        hh.setStretchLastSection(False)
        for i in range(len(headers)):
            hh.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)


    def _report_monthly_revenue(self):
        data = self._svc.monthly_revenue(12)
        currencies = set()
        for row in data:
            currencies.update(k for k in row if k != "month")
        currencies = sorted(currencies)

        headers = ["Ay"] + [f"Ciro ({c})" for c in currencies]
        self._setup_columns(headers)
        self._table.setRowCount(len(data))

        total_revenue = 0
        for r, row in enumerate(data):
            self._table.setItem(r, 0, QTableWidgetItem(row["month"]))
            for ci, cur in enumerate(currencies):
                val = row.get(cur, 0)
                total_revenue += val
                sym = SYM_MAP.get(cur, cur)
                text = fmt_money(val, sym) if val > 0 else "-"
                item = QTableWidgetItem(text)
                item.setTextAlignment(
                    Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self._table.setItem(r, ci + 1, item)

        if total_revenue > 0:
            self._summary_label.setText(
                f"Son 12 ayın aylık ciro raporu  |  "
                f"{len(currencies)} para birimi  |  "
                f"Toplam: {fmt_number(total_revenue)}")
        else:
            self._summary_label.setText("Son 12 ayda teklif bulunmuyor.")

    def _report_customer_ranking(self):
        data = self._svc.customer_ranking(20)
        headers = ["Firma", "Para Birimi", "Teklif", "Onay", "Toplam Ciro", "Son Teklif"]
        self._setup_columns(headers)
        self._table.setRowCount(len(data))

        for r, row in enumerate(data):
            self._table.setItem(r, 0, QTableWidgetItem(row["company_name"]))
            self._table.setItem(r, 1, QTableWidgetItem(row.get("currency", "")))
            cnt = QTableWidgetItem(str(row["offer_count"]))
            cnt.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self._table.setItem(r, 2, cnt)
            appr = QTableWidgetItem(str(row.get("approved", 0)))
            appr.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self._table.setItem(r, 3, appr)
            sym = SYM_MAP.get(row.get("currency", ""), "")
            rev = QTableWidgetItem(fmt_money(row['total_revenue'], sym))
            rev.setTextAlignment(
                Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self._table.setItem(r, 4, rev)
            self._table.setItem(r, 5, QTableWidgetItem(
                to_display_date(row.get("last_date", ""))))

        self._summary_label.setText(f"En yüksek cirolu {len(data)} müşteri (iptal hariç)")

    def _report_product_ranking(self):
        data = self._svc.product_ranking(20)
        headers = ["Ürün Kodu", "Ürün Adı", "Teklif", "Miktar", "Tutar"]
        self._setup_columns(headers)
        self._table.setRowCount(len(data))

        for r, row in enumerate(data):
            self._table.setItem(r, 0, QTableWidgetItem(row.get("product_code", "")))
            self._table.setItem(r, 1, QTableWidgetItem(row.get("product_name", "")))
            cnt = QTableWidgetItem(str(row["offer_count"]))
            cnt.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self._table.setItem(r, 2, cnt)
            qty = QTableWidgetItem(fmt_number(row['total_qty'], 0))
            qty.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self._table.setItem(r, 3, qty)
            rev = QTableWidgetItem(fmt_number(row['total_revenue']))
            rev.setTextAlignment(
                Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self._table.setItem(r, 4, rev)

        self._summary_label.setText(f"En çok teklif edilen {len(data)} ürün (iptal hariç)")

    def _report_conversion(self):
        data = self._svc.conversion_rate()
        avg = self._svc.average_offer_amount()

        headers = ["Metrik", "Değer"]
        self._setup_columns(headers)

        rate = data["rate"]
        rate_str = f"{rate:g}" if rate == int(rate) else f"{rate:.1f}"

        rows = [
            ("Toplam Teklif Sayısı", str(data["total"])),
            ("Onaylanan Teklif", str(data["approved"])),
            ("Beklemede", str(data["pending"])),
            ("İptal Edilen", str(data["cancelled"])),
            ("", ""),
            ("Dönüşüm Oranı", f"%{rate_str}"),
        ]
        for cur, info in sorted(avg.items()):
            sym = SYM_MAP.get(cur, cur)
            rows.append((
                f"Ortalama Teklif Tutarı ({cur})",
                f"{fmt_money(info['avg'], sym)}  ({info['count']} teklif)"))

        self._table.setRowCount(len(rows))
        for r, (label, value) in enumerate(rows):
            lbl_item = QTableWidgetItem(label)
            val_item = QTableWidgetItem(value)
            if not label:
                lbl_item.setFlags(lbl_item.flags() & ~Qt.ItemFlag.ItemIsSelectable)
                val_item.setFlags(val_item.flags() & ~Qt.ItemFlag.ItemIsSelectable)
            val_item.setTextAlignment(
                Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self._table.setItem(r, 0, lbl_item)
            self._table.setItem(r, 1, val_item)

        self._summary_label.setText(
            f"Dönüşüm oranı: %{rate_str} "
            f"({data['approved']} onay / {data['total']} toplam)")

    def _export(self):
        import datetime
        default = f"rapor_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "Raporu Kaydet", default, "Excel (*.xlsx)")
        if not path:
            return
        if self._table.rowCount() == 0:
            QMessageBox.information(self, "Bilgi", "Dışa aktarılacak veri yok.\nÖnce bir rapor oluşturun.")
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = Workbook()
            ws = wb.active
            ws.title = self._report_combo.currentText()

            header_fill = PatternFill("solid", fgColor="1E4D8C")
            header_font = Font(bold=True, color="FFFFFF", size=11)

            for col in range(self._table.columnCount()):
                hdr_item = self._table.horizontalHeaderItem(col)
                cell = ws.cell(1, col + 1, hdr_item.text() if hdr_item else "")
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row in range(self._table.rowCount()):
                for col in range(self._table.columnCount()):
                    item = self._table.item(row, col)
                    ws.cell(row + 2, col + 1, item.text() if item else "")

            wb.save(path)
            QMessageBox.information(self, "Kaydedildi", f"Rapor kaydedildi:\n{path}")
        except Exception as e:
            QMessageBox.warning(self, "Hata", f"Export hatası:\n{e}")
