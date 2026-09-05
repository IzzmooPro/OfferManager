"""
Excel / CSV veri aktarımı — penceresiz akış.

İçe aktarma:  Dosya → İçeri Aktar → dosya seç → özet onayı → aktar.
Dışa aktarma: Dosya → Dışarı Aktar → kayıt yeri seç → yaz.

Şablon sütun isimleri esnek eşleşir (Türkçe/İngilizce, büyük-küçük harf).
Dışa aktarılan dosyalar içe aktarmayla birebir uyumludur (roundtrip).
"""
import logging, csv, io, re
from pathlib import Path
from PySide6.QtWidgets import QFileDialog, QMessageBox, QCheckBox
from core.constants import normalize_currency
from ui.utils import operation_error as op_hata
from ui.utils import operation_error_dialog as hata_diyalogu

logger = logging.getLogger("excel_import")

# ── Sabit kullanıcı metinleri (invariant 18) ─────────────────────────────────
# Teknik neden ASLA metne girmez; yalnız güvenli loga yazılır.
DOSYA_OKUMA_HATASI = (
    "Dosya okunamadı. Dosyanın başka bir programda açık olmadığından ve "
    "biçiminin doğru olduğundan emin olup yeniden deneyin.")
SATIR_YAZILAMADI = "Satır {sira}: kaydedilemedi."
TEKLIF_YAZILAMADI = "{sira}. teklif kaydedilemedi."
KATEGORI_UYARISI = (
    "Bazı kategoriler oluşturulamadı; ilgili ürünler kategorisiz kaydedildi.")
ASAMA_BASARISIZ = {
    "customers": "Müşteri: aktarılamadı",
    "products":  "Ürün: aktarılamadı",
    "offers":    "Teklif: aktarılamadı",
}

# Sütun eşleştirme haritaları (olası başlık → alan adı)
CUSTOMER_MAP = {
    "firma adı": "company_name",  "firma": "company_name",
    "şirket adı": "company_name", "şirket": "company_name",
    "company": "company_name",    "company name": "company_name",
    "ilgili kişi": "contact_person", "ilgili": "contact_person",
    "kişi": "contact_person",     "contact": "contact_person",
    "adres": "address",           "address": "address",
    "telefon": "phone",           "tel": "phone",
    "phone": "phone",             "gsm": "phone",
    "e-posta": "email",           "eposta": "email",
    "email": "email",             "mail": "email",
    "not": "notes",               "notes": "notes",
    "müşteri notu": "notes",
}
PRODUCT_MAP = {
    "ürün kodu": "product_code",  "kod": "product_code",
    "code": "product_code",       "product code": "product_code",
    "ürün adı": "product_name",   "ürün": "product_name",
    "ad": "product_name",         "name": "product_name",
    "açıklama": "description",    "aciklama": "description",
    "description": "description", "detay": "description",
    "fiyat": "price",             "price": "price",
    "birim fiyat": "price",       "unit price": "price",
    "para birimi": "currency",    "currency": "currency",
    "döviz": "currency",
    "stok": "stock",              "stock": "stock",
    "miktar": "stock",            "quantity": "stock",
    "birim": "unit",              "unit": "unit",
    "kategori": "category",       "category": "category",
}


class _ImportProgress:
    """İçe aktarma sırasında gerçek ilerlemeyi gösteren modal pencere.

    Yüzde satır sayımından hesaplanır (sahte animasyon değil). Çağrılabilir:
    `prog(current, total)` → çubuğu günceller. İşlemler tek transaction'da hızlı
    olduğundan iptal butonu yok — pencere yalnız durumu doğru yansıtır.
    """

    def __init__(self, parent, label: str):
        from PySide6.QtWidgets import QProgressDialog
        from PySide6.QtCore import Qt
        dlg = QProgressDialog(label, "", 0, 100, parent)
        dlg.setWindowTitle("İçe Aktarma")
        dlg.setWindowModality(Qt.WindowModality.WindowModal)
        dlg.setCancelButton(None)          # yarım kalırsa veri tutarsız olmasın
        dlg.setMinimumWidth(360)
        dlg.setMinimumDuration(0)          # hemen görün
        dlg.setAutoClose(False)
        dlg.setAutoReset(False)
        dlg.setValue(0)
        self._dlg = dlg
        self._last = -1

    def set_label(self, text: str):
        self._dlg.setLabelText(text)
        self._dlg.setValue(0)
        self._last = -1
        from PySide6.QtWidgets import QApplication
        QApplication.processEvents()

    def __call__(self, current: int, total: int):
        pct = int(current * 100 / total) if total else 100
        pct = min(max(pct, 0), 100)
        if pct != self._last:               # yalnız değişince yeniden çiz
            self._last = pct
            self._dlg.setValue(pct)
            from PySide6.QtWidgets import QApplication
            QApplication.processEvents()

    def close(self):
        self._dlg.close()


def _norm(s: str) -> str:
    """Başlığı eşleştirme için normalize et — Türkçe İ/I harflerine dikkat.

    Python'da "İ".lower() sonucu "i" + birleşik nokta (U+0307) çıkar; bu
    yüzden "İlgili Kişi" gibi başlıklar haritayla eşleşemiyordu. Türkçe
    kurala göre önce İ→i ve I→ı çevrilir, sonra küçültülür.
    """
    s = s.strip().replace("İ", "i").replace("I", "ı")
    return s.lower().replace("̇", "").replace("_", " ")


_CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1254", "latin-1")
_CSV_DELIMITERS = ",;\t"
# Kullanıcı sayfa seçimini iptal ettiğinde `_read_file`'ın döndürdüğü işaret:
# hata mesajı GÖSTERİLMEZ, akış sessizce sonlanır.
SAYFA_SECIMI_IPTAL = "__sayfa_secimi_iptal__"

_CSV_OKUMA_HATASI = (
    "Dosya okunamadı — içeriği bozuk ya da CSV değil.\n"
    "Dosyayı Excel'de açıp 'CSV UTF-8' biçiminde yeniden kaydedip deneyin.")


def _workbook_kapat(wb) -> None:
    """Workbook'u kapatır; kapatma hatasını DIŞARI SIZDIRMAZ.

    Kapatma, okumadan AYRI bir aşamadır: başarılı bir okumayı "dosya
    okunamadı"ya çeviremez, sayfa seçimi/iptal sonucunu değiştiremez ve asıl
    okuma hatasını maskeleyemez. Hata sabit işlem adıyla tam bir kez güvenli
    loglanır (asıl istisna ayrıca kendi yerinde loglanır).
    """
    if wb is None:
        return
    try:
        wb.close()
    except Exception as exc:                                   # noqa: BLE001
        op_hata.logla(exc, "Calisma kitabi kapat")


def _ikili_gorunumlu(ham: bytes) -> bool:
    """İçerik metin mi? NUL baytı veya yoğun kontrol karakteri → metin değil.

    latin-1 HER bayt dizisini hatasız çözdüğünden, bu kontrol olmadan ikili
    dosyalar anlamsız satırlara dönüşüp "okundu" sanılıyordu.
    """
    ornek = ham[:4096]
    if b"\x00" in ornek:
        return True
    if not ornek:
        return False
    bozuk = sum(1 for b in ornek if b < 9 or 13 < b < 32)
    return bozuk / len(ornek) > 0.30


def _csv_ayraci(text: str, kaynak: str) -> str:
    """Ayıracı belirler; Sniffer başarısızsa güvenli yedeğe düşer.

    Tek sütunlu veya sıra dışı ama OKUNABİLİR dosyalar reddedilmez; zorunlu
    sütun denetimi sonraki doğrulama katmanının işidir. Sniffer hatası
    sessizce yutulmaz, log'a yazılır.
    """
    try:
        return csv.Sniffer().sniff(
            text[:2048], delimiters=_CSV_DELIMITERS).delimiter
    except csv.Error as exc:
        ilk_satir = next((s for s in text.splitlines() if s.strip()), "")
        ayrac = max(_CSV_DELIMITERS, key=ilk_satir.count)
        if ilk_satir.count(ayrac) == 0:
            ayrac = ","
        # BEKLENEN fallback: hata değil. Ham istisna ve dosya adı (kullanıcı
        # verisi taşıyabilir) loga yazılmaz; yalnız seçilen yedek ayıraç.
        logger.info("CSV ayıracı belirlenemedi (%s) — yedek ayıraç %r "
                    "kullanılıyor", type(exc).__name__, ayrac)
        return ayrac


def _read_csv(path: Path) -> tuple[list, str]:
    """CSV'yi bilinen kodlamalarla okur → (satırlar, kullanıcı hatası).

    GERÇEKTEN boş dosya ile OKUNAMAYAN dosya ayrılır: tüm denemeler
    başarısız olursa hata alanı boş bırakılmaz (eskiden kullanıcı gerçek
    neden yerine "Dosyada veri bulunamadı." görüyordu). Teknik ayrıntı
    log'a yazılır, kullanıcıya kısa bir açıklama döner.
    """
    try:
        ham = path.read_bytes()
    except OSError as exc:
        op_hata.logla(exc, "CSV oku")
        return [], _CSV_OKUMA_HATASI

    if not ham.strip():
        return [], ""                      # gerçekten boş dosya
    if _ikili_gorunumlu(ham):
        # İstisna yok; dosya adı kullanıcı verisi taşıyabildiği için yazılmaz.
        logger.warning("CSV metin içermiyor: ikili/bozuk içerik")
        return [], _CSV_OKUMA_HATASI

    son_hata = None
    for enc in _CSV_ENCODINGS:
        try:
            text = ham.decode(enc)
        except (UnicodeDecodeError, LookupError) as exc:
            son_hata = exc
            # Beklenen deneme: ham istisna ve dosya adı yazılmaz.
            logger.debug("CSV kodlama denemesi başarısız (%s): %s",
                         enc, type(exc).__name__)
            continue
        try:
            reader = csv.DictReader(io.StringIO(text),
                                    delimiter=_csv_ayraci(text, path.name))
            return [dict(r) for r in reader], ""
        except csv.Error as exc:
            son_hata = exc
            logger.debug("CSV ayrıştırma başarısız (%s): %s",
                         enc, type(exc).__name__)

    # Tüm denemeler tükendi: SON anlamlı teknik neden tam bir kez güvenli
    # loglanır (kaybolmaz), kullanıcıya sabit metin döner.
    if son_hata is not None:
        op_hata.logla(son_hata, "CSV ayristir")
    else:
        logger.warning("CSV ayrıştırılamadı: bilinen kodlama bulunamadı")
    return [], _CSV_OKUMA_HATASI


def _sayfa_alanlari(headers, import_type: str) -> set:
    """Başlık satırındaki TANINAN alan adları — mevcut eşleme mekanizmasıyla."""
    col_map = CUSTOMER_MAP if import_type == "customers" else PRODUCT_MAP
    return {col_map.get(_norm(h)) for h in headers} - {None}


def _sayfa_adaylari(wb, import_type: str):
    """(adaylar, gizli_dolu_sayfa_var_mi) döndürür.

    Aday ölçütü: GÖRÜNÜR sayfa + zorunlu başlıkların tamamı + en az bir gerçek
    veri satırı. Yalnız başlık, boş ve açıklama sayfaları aday olmaz; gizli
    sayfalar hiç değerlendirilmez (otomatik alınmaz, listelenmez).
    """
    gerekli = ({"company_name"} if import_type == "customers"
               else {"product_code", "product_name"})
    adaylar, gizli_dolu = [], False
    for ws in wb.worksheets:
        gorunur = getattr(ws, "sheet_state", "visible") == "visible"
        ilk = None
        veri_var = False
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx == 0:
                ilk = [str(c or "").strip() for c in row]
                continue
            if any(c is not None and str(c).strip() for c in row):
                veri_var = True
                break
        if not ilk or not veri_var:
            continue
        if not gerekli.issubset(_sayfa_alanlari(ilk, import_type)):
            continue
        if gorunur:
            adaylar.append((ws.title, max(0, (ws.max_row or 1) - 1)))
        else:
            gizli_dolu = True
    return adaylar, gizli_dolu


def _gizli_notu(gizli_dolu: bool) -> str:
    return "\nGizli çalışma sayfaları içe aktarılmaz." if gizli_dolu else ""


def _sayfa_sordur(parent, adaylar, aktif_ad, gizli_dolu):
    """Birden çok aday varsa kullanıcıya sordurur. İptalde None döner.

    Satır sayısı YAKLAŞIKTIR: `ws.max_row` üzerinden hesaplanır, kesin sayım
    için büyük sayfa ikinci kez taranmaz. Seçim, görünen metinden ad
    ayrıştırılarak değil, metin→sayfa adı eşlemesiyle çözülür.
    """
    from PySide6.QtWidgets import QInputDialog
    eslesme = {}
    secenekler = []
    for ad, n in adaylar:
        metin = f"{ad} — yaklaşık {n} satır"
        eslesme[metin] = ad
        secenekler.append(metin)
    varsayilan = next((i for i, (ad, _) in enumerate(adaylar)
                       if ad == aktif_ad), 0)
    mesaj = ("Bu dosyada birden fazla uygun veri sayfası var.\n"
             "Yalnız seçtiğiniz sayfa içe aktarılacaktır."
             + _gizli_notu(gizli_dolu))
    secim, ok = QInputDialog.getItem(parent, "Çalışma Sayfası Seç", mesaj,
                                     secenekler, varsayilan, False)
    if not ok:
        return None
    # Beklenmedik bir dönüş gelirse ham ValueError sızdırma; güvenli iptal.
    return eslesme.get(secim)


def _aday_yok_mesaji(wb, import_type: str, gizli_dolu: bool) -> str:
    tur = "müşteri" if import_type == "customers" else "ürün"
    gorunur = sum(1 for ws in wb.worksheets
                  if getattr(ws, "sheet_state", "visible") == "visible")
    mesaj = (f"Bu dosyada içe aktarılabilir bir {tur} sayfası bulunamadı.\n"
             f"Görünür çalışma sayfası sayısı: {gorunur}.\n"
             "Sayfanın ilk satırında beklenen başlıklar ve altında en az bir "
             "veri satırı olmalıdır.")
    if gizli_dolu:
        mesaj += ("\n\nDosyada uygun görünen GİZLİ sayfa var; gizli çalışma "
                  "sayfaları içe aktarılmaz. Aktarmak için sayfayı Excel'de "
                  "görünür yapın.")
    return mesaj


def _sayfa_sec_onceden(path: str, import_type: str, parent) -> tuple:
    """Sayfa seçimini İLERLEME PENCERESİ AÇILMADAN ÖNCE tamamlar.

    Neden ayrı adım: `run_import_flow` ilerleme penceresini modal açtığı için,
    okuma sırasında sorulan sayfa sorusu Windows'ta DEVRE DIŞI kalıyordu —
    kullanıcı soruyu görüyor ama yanıtlayamıyordu (akış süresiz kilitleniyordu).

    Döner: `(secilen_sayfa, hata)`.
      * XLSX değilse / açılamıyorsa / aday sayısı 1'den fazla değilse
        `(None, "")` döner; asıl karar ve hata metni `_read_file`'a kalır
        (davranış değişmez, ikinci bir hata kaynağı oluşmaz).
      * Kullanıcı iptal ederse `(None, SAYFA_SECIMI_IPTAL)`.
    """
    if Path(path).suffix.lower() not in (".xlsx", ".xls", ".xlsm"):
        return None, ""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None, ""          # hata mesajını _read_file üretir
    try:
        adaylar, gizli_dolu = _sayfa_adaylari(wb, import_type)
        if len(adaylar) < 2:
            return None, ""      # soru sorulmaz; tek aday/aday yok yolu aynı
        secilen = _sayfa_sordur(parent, adaylar, wb.active.title, gizli_dolu)
        if secilen is None:
            return None, SAYFA_SECIMI_IPTAL
        return secilen, ""
    finally:
        _workbook_kapat(wb)


def _read_file(path: str, progress=None, import_type: str = "products",
               parent=None, secilen_sayfa: str = None) -> tuple[list, str]:
    """Dosyayı okur, (rows, error) döndürür. rows = list of dicts.

    XLSX'te yalnız SEÇİLEN çalışma sayfası okunur; sayfalar otomatik
    BİRLEŞTİRİLMEZ. Tek geçerli görünür sayfa varsa otomatik seçilir; birden
    fazlaysa kullanıcıya sorulur. Kullanıcı iptal ederse hata alanı
    `SAYFA_SECIMI_IPTAL` olur ve çağıran sessizce çıkar.

    `secilen_sayfa` verilirse (bkz. `_sayfa_sec_onceden`) soru TEKRAR sorulmaz;
    verilmezse davranış eskisiyle birebir aynıdır.

    `progress(current, total)` verilirse xlsx okurken satır ilerlemesi bildirilir.
    """
    p = Path(path)
    ext = p.suffix.lower()
    rows = []
    try:
        if ext in (".xlsx", ".xls", ".xlsm"):
            try:
                import openpyxl
            except ImportError:
                return [], ("openpyxl kütüphanesi bulunamadı.\n"
                            "Lütfen dosyayı CSV olarak kaydedin veya\n"
                            "komut satırında: pip install openpyxl")
            # Workbook TEK kez açılır: aday tespiti ve okuma aynı nesneden.
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                adaylar, gizli_dolu = _sayfa_adaylari(wb, import_type)
                if not adaylar:
                    return [], _aday_yok_mesaji(wb, import_type, gizli_dolu)
                if secilen_sayfa in [a[0] for a in adaylar]:
                    secilen = secilen_sayfa      # soru zaten sorulmuş
                elif len(adaylar) == 1:
                    secilen = adaylar[0][0]
                else:
                    secilen = _sayfa_sordur(parent, adaylar,
                                            wb.active.title, gizli_dolu)
                    if secilen is None:
                        return [], SAYFA_SECIMI_IPTAL
                ws = wb[secilen]
                # Satır satır akış — büyük dosyada belleğe komple almak yerine
                # okurken ilerleme bildir (toplam ws.max_row'dan tahmin edilir).
                total = ws.max_row or 0
                headers = None
                for idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if idx == 0:
                        headers = [str(c or "").strip() for c in row]
                        continue
                    if all(c is None for c in row): continue
                    satir = {headers[i]: (str(v) if v is not None else "")
                             for i, v in enumerate(row) if i < len(headers)}
                    satir["_source_sheet"] = secilen
                    rows.append(satir)
                    if progress and total and (idx & 0x3FF) == 0:
                        progress(idx, total)
                if headers is None:
                    return [], "Dosya boş."
            finally:
                _workbook_kapat(wb)
        elif ext == ".csv":
            return _read_csv(p)
        else:
            return [], f"Desteklenmeyen dosya türü: {ext}\nDesteklenen: .xlsx, .csv"
    except Exception as exc:                                   # noqa: BLE001
        # Ham istisna kullanıcıya GİTMEZ; teknik neden tam bir kez güvenli
        # loglanır. `_sayfa_sec_onceden` aynı hatayı bilerek loglamaz.
        op_hata.logla(exc, "Dosya oku")
        return [], DOSYA_OKUMA_HATASI
    return rows, ""


def _map_row(row: dict, col_map: dict) -> dict:
    """Ham satırı alan adlarına çevirir."""
    result = {}
    for raw_key, value in row.items():
        norm_key = _norm(raw_key)
        field = col_map.get(norm_key)
        if field:
            result[field] = value.strip() if isinstance(value, str) else (
                value if value is not None else "")
    return result


def _parse_number(value, default: float = 0.0) -> float:
    """Türkçe ve uluslararası binlik/ondalık biçimlerini güvenli ayrıştır."""
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[^0-9,\.\-+]", "", str(value).strip())
    if not text:
        return default
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    elif text.count(".") > 1:
        text = text.replace(".", "")
    try:
        return float(text)
    except ValueError:
        return default


# ── Doğrulama ─────────────────────────────────────────────────────────────────

def _validate_rows(import_type: str, raw_rows: list,
                   progress=None) -> tuple[list, list, list]:
    """Ham satırları eşleştirir ve (geçerli, mükerrer, hatalı) olarak ayırır.

    Mükerrer kontrolü mevcut anahtarları TEK sorguda belleğe alıp orada yapılır
    (satır başına ayrı DB sorgusu on binlerce satırda ~1600x daha yavaştı).
    `progress(current, total)` verilirse aşama ilerlemesi bildirilir.
    """
    from database.db_manager import get_db
    db = get_db()
    col_map = CUSTOMER_MAP if import_type == "customers" else PRODUCT_MAP
    required = ("company_name",) if import_type == "customers" else (
        "product_code", "product_name")

    # Mevcut anahtarlar → id (TEK sorgu). Müşteri: firma adı (harfi harfine),
    # Ürün: normalize_code() ile NFKC + casefold; en düşük id deterministik
    # olarak kazanır.
    if import_type == "customers":
        existing = {(r["company_name"] or "").strip(): r["id"]
                    for r in db.fetchall("SELECT id, company_name FROM customers")}
    else:
        # Servisle AYNI anahtar (NFKC + casefold) ve ORDER BY id ile
        # deterministik eşleme: eski DB'de çakışma varsa en düşük id kazanır.
        from services.product_service import normalize_code
        existing = {}
        for r in db.fetchall(
                "SELECT id, product_code FROM products ORDER BY id"):
            anahtar = normalize_code(r["product_code"])
            if not anahtar:
                continue
            if anahtar in existing:
                logger.warning(
                    "Ürün kodu %r birden fazla kayıtta; en düşük id (%s) "
                    "kullanılıyor (id=%s atlandı).",
                    anahtar, existing[anahtar], r["id"])
                continue
            existing[anahtar] = r["id"]

    valid, duplicates, invalid = [], [], []
    # Aynı dosyada tekrar eden anahtarlar. Müşteri anahtarı DB eşleşmesiyle
    # BİREBİR aynıdır: (company_name or "").strip() — casefold/NFKC UYGULANMAZ,
    # yani 'Acme' ile 'ACME' farklı müşteri sayılır (aynı isimli farklı gerçek
    # müşteriler olabilir). Ürün anahtarı O6'daki normalize_code'dur.
    # Sayaç sözleşmesi: dosya içi tekrar `invalid`e YALNIZ BİR KEZ yazılır;
    # `duplicate` (DB'de zaten var) ile örtüşmez.
    dosyadaki = set()
    total = len(raw_rows)
    for i, raw in enumerate(raw_rows):
        r = _map_row(raw, col_map)
        # Kaynak sayfa adı `_map_row` sütun eşlemesinden geçmez; hata
        # mesajlarında gösterilebilmesi için elle taşınır.
        sayfa = raw.get("_source_sheet") if isinstance(raw, dict) else None
        if sayfa:
            r["_source_sheet"] = sayfa
        missing = [k for k in required if not str(r.get(k, "")).strip()]
        if missing:
            r["_error"] = f"Zorunlu alan eksik: {', '.join(missing)}"
            invalid.append(r)
        else:
            urun = import_type != "customers"
            if urun:
                from services.product_service import normalize_code
                key = normalize_code(r.get("product_code", ""))
            else:
                key = r.get("company_name", "").strip()
            # Dosya içi tekrar: ilk GEÇERLİ satır anahtarı sahiplenir,
            # sonrakiler işlenmez. Zorunlu alanı geçemeyen satırlar buraya hiç
            # ulaşmadığı için adı/kodu sahiplenmez. Sessizce ikinci bir kayıt
            # oluşturmak veya aynı hedefi iki kez güncellemek yerine bildirilir.
            if key in dosyadaki:
                r["_error"] = (
                    "Bu ürün kodu dosyada birden fazla kez var; "
                    "yalnız ilk satır işlendi" if urun else
                    "Bu firma adı dosyada birden fazla kez var; "
                    "yalnız ilk satır işlendi")
                invalid.append(r)
            else:
                dosyadaki.add(key)
                ex_id = existing.get(key)
                if ex_id is not None:
                    r["_duplicate"] = True
                    r["_existing_id"] = ex_id
                    duplicates.append(r)
                else:
                    valid.append(r)
        if progress and (i & 0x3FF) == 0:   # ~her 1024 satırda bir güncelle
            progress(i + 1, total)
    if progress:
        progress(total, total)
    # Hata metnine kaynak sayfa adını okunur biçimde ekle (iç alan adı değil).
    for r in invalid:
        sayfa = r.get("_source_sheet")
        if sayfa and r.get("_error"):
            r["_error"] = f"{r['_error']} (sayfa: {sayfa})"
    return valid, duplicates, invalid


# ── Veritabanına yazma ────────────────────────────────────────────────────────

def _perform_import(import_type: str, rows_to_process: list,
                    update_dups: bool, progress=None,
                    stage_state: dict = None) -> tuple[int, int, int, list]:
    """Satırları veritabanına yazar. (eklenen, güncellenen, atlanan, hatalar)

    `progress(current, total)` verilirse kaydetme ilerlemesi bildirilir.

    `stage_state` verilirse, ürün transaction'ından ÖNCE tamamlanan bağımsız
    aşamalar buraya işlenir (`kategori_yazildi` = kaç yeni kategori yazıldı).
    Böylece transaction düşse bile çağıran, DB'nin gerçekten değiştiğini
    bilir. İçine YALNIZ sayı/boolean yazılır — kategori adı ya da başka
    kullanıcı verisi taşımaz (invariant 18).
    """
    from database.db_manager import get_db
    db = get_db()
    added = updated = skipped = 0
    errors = []
    total = len(rows_to_process)

    # Tüm satırlar TEK transaction'da yazılır — satır başına ayrı commit
    # (her biri diske fsync) yerine tek diske-yazma. Binlerce satırda ~1000x
    # hız farkı. Satır-içi try/except korunur: bir bozuk satır tüm aktarımı
    # iptal etmez, sadece o satır atlanır (başarısız INSERT atomik geri alınır,
    # transaction açık kalır).
    if import_type == "customers":
        with db.transaction() as conn:
            for i, row in enumerate(rows_to_process):
                company = row.get("company_name", "").strip()
                if not company:
                    skipped += 1
                    continue
                is_dup = row.get("_duplicate", False)
                try:
                    if is_dup and update_dups:
                        conn.execute(
                            "UPDATE customers SET contact_person=?, address=?, "
                            "phone=?, email=?, notes=? WHERE id=?",
                            (row.get("contact_person", ""), row.get("address", ""),
                             row.get("phone", ""), row.get("email", ""),
                             row.get("notes", ""), row["_existing_id"]))
                        updated += 1
                    else:
                        conn.execute(
                            "INSERT INTO customers (company_name,contact_person,"
                            "address,phone,email,notes) VALUES (?,?,?,?,?,?)",
                            (company, row.get("contact_person", ""),
                             row.get("address", ""), row.get("phone", ""),
                             row.get("email", ""), row.get("notes", "")))
                        added += 1
                except Exception as exc:                       # noqa: BLE001
                    # Firma adı ve ham istisna kullanıcıya GİTMEZ; güvenli
                    # kayit_id olarak yalnız satır SIRASI kullanılır.
                    op_hata.logla(exc, "Musteri satiri yaz", kayit_id=i + 1)
                    errors.append(SATIR_YAZILAMADI.format(sira=i + 1))
                    skipped += 1
                if progress and (i & 0xFF) == 0:
                    progress(i + 1, total)
    else:
        # Kategoriler transaction'dan ÖNCE çözülür (yoksa oluşturulur) — böylece
        # tek yazıcı kuralı gereği iç içe yazma kilidi (SQLITE_BUSY) oluşmaz.
        from services.category_service import CategoryService
        from models.category import Category
        cat_svc = CategoryService()
        cat_cache = {c.name.strip().casefold(): c.id for c in cat_svc.get_all()}
        kategori_hatasi = False
        kategori_yazildi = 0
        for row in rows_to_process:
            nm = (row.get("category", "") or "").strip()
            if nm and nm.casefold() not in cat_cache:
                try:
                    cat_cache[nm.casefold()] = cat_svc.add(Category(name=nm))
                    kategori_yazildi += 1
                except Exception as exc:                       # noqa: BLE001
                    # Başarısızlık ÖNBELLEĞE alınır: AYNI kategori yeniden
                    # denenmez. FARKLI kategoriler ayrı istisnalardır ve her
                    # biri tam bir kez loglanır; kullanıcıya ise tek toplu
                    # uyarı gösterilir. Kategori adı hiçbir yere yazılmaz.
                    cat_cache[nm.casefold()] = None
                    kategori_hatasi = True
                    op_hata.logla(exc, "Kategori olustur")
        if stage_state is not None:
            # Kategoriler ürün transaction'ından ÖNCE ve ONDAN BAĞIMSIZ
            # yazılır: transaction düşse bile bu değişiklik DB'de kalır.
            stage_state["kategori_yazildi"] = kategori_yazildi
        if kategori_hatasi:
            # Ürünler kategorisiz KAYDEDİLMEYE devam eder; bu tamamlanmış
            # ürün kaydını inkâr etmeyen tek satırlık kısmi sonuçtur (18b).
            errors.append(KATEGORI_UYARISI)

        def _category_id(raw_name: str):
            nm = (raw_name or "").strip()
            return cat_cache.get(nm.casefold()) if nm else None

        with db.transaction() as conn:
            for i, row in enumerate(rows_to_process):
                code = row.get("product_code", "").strip()
                name = row.get("product_name", "").strip()
                if not code or not name:
                    skipped += 1
                    continue
                price = _parse_number(row.get("price", 0))
                stock = _parse_number(row.get("stock", 0))
                # TRY/₺ gibi yaygın yazımlar sistemin "TL" koduna eşlenir;
                # tanınmayan/boş → EUR (bkz. core.constants.normalize_currency).
                currency = normalize_currency(row.get("currency", ""), default="EUR")
                unit = row.get("unit", "Adet") or "Adet"
                is_dup = row.get("_duplicate", False)
                try:
                    cat_id = _category_id(row.get("category", ""))
                    if is_dup and update_dups:
                        conn.execute(
                            "UPDATE products SET product_name=?, description=?, "
                            "price=?, currency=?, stock=?, unit=?, category_id=? "
                            "WHERE id=?",
                            (name, row.get("description", ""), price, currency,
                             stock, unit, cat_id, row["_existing_id"]))
                        updated += 1
                    else:
                        conn.execute(
                            "INSERT INTO products (product_code,product_name,"
                            "description,price,currency,stock,unit,category_id) "
                            "VALUES (?,?,?,?,?,?,?,?)",
                            (code, name, row.get("description", ""),
                             price, currency, stock, unit, cat_id))
                        added += 1
                except Exception as exc:                       # noqa: BLE001
                    # Ürün kodu ve ham istisna kullanıcıya GİTMEZ.
                    op_hata.logla(exc, "Urun satiri yaz", kayit_id=i + 1)
                    errors.append(SATIR_YAZILAMADI.format(sira=i + 1))
                    skipped += 1
                if progress and (i & 0xFF) == 0:
                    progress(i + 1, total)

    if progress:
        progress(total, total)
    return added, updated, skipped, errors


# ── Kullanıcı akışları ────────────────────────────────────────────────────────

def run_import_flow(parent, import_type: str) -> bool:
    """Dosya seç → özet onayı → aktar. Aktarım yapıldıysa True döner."""
    label = {"customers": "müşteri", "products": "ürün",
             "offers": "teklif"}[import_type]
    path, _ = QFileDialog.getOpenFileName(
        parent, f"{label.capitalize()} Verisi İçe Aktar", "",
        "Excel & CSV (*.xlsx *.xls *.xlsm *.csv);;Tüm Dosyalar (*)")
    if not path:
        return False

    # Sayfa sorusu İLERLEME PENCERESİNDEN ÖNCE sorulur: modal ilerleme
    # penceresi açıkken sorulan soru Windows'ta devre dışı kalıyordu.
    secilen_sayfa, sayfa_hatasi = _sayfa_sec_onceden(path, import_type, parent)
    if sayfa_hatasi == SAYFA_SECIMI_IPTAL:
        return False                       # sessiz iptal: pencere/DB yazımı yok

    prog = _ImportProgress(parent, "Dosya okunuyor…")
    raw_rows, err = _read_file(path, progress=prog, import_type=import_type,
                               parent=parent, secilen_sayfa=secilen_sayfa)
    if err == SAYFA_SECIMI_IPTAL:
        # Kullanıcı sayfa seçimini iptal etti: doğrulama, DB yazımı ve sonuç
        # mesajı YOK; işlem temiz biçimde iptal sayılır.
        prog.close()
        return False
    if err:
        prog.close()
        QMessageBox.warning(parent, "Dosya Hatası", err)
        return False
    if not raw_rows:
        prog.close()
        QMessageBox.warning(parent, "Boş Dosya", "Dosyada veri bulunamadı.")
        return False

    if import_type == "offers":
        # Teklifler satır-bazlı değil grup-bazlı doğrulanır (kalemli format).
        # Okuma ilerleme penceresi BURADA kapanır; alt akış kendi penceresini
        # yönetir. Alt akıştan kaçan beklenmeyen istisna dışarı sızmaz.
        prog.close()
        try:
            return _run_offer_import_flow(parent, path, raw_rows)
        except Exception as exc:                               # noqa: BLE001
            hata_diyalogu.hata_goster(parent, "Hata", exc, "Teklif", "aktar")
            return False

    # ── AŞAMA: DOĞRULAMA — henüz DB'ye hiçbir şey yazılmadı ──────────────
    prog.set_label("Kayıtlar denetleniyor…")
    try:
        valid, duplicates, invalid = _validate_rows(
            import_type, raw_rows, progress=prog)
    except Exception as exc:                                   # noqa: BLE001
        prog.close()
        hata_diyalogu.hata_goster(parent, "Hata", exc, "İçe aktarma", "denetle")
        return False
    prog.close()
    if not valid and not duplicates:
        msg = f"Dosyada aktarılabilir {label} kaydı bulunamadı."
        if invalid:
            msg += "\n\nHatalı satırlar:\n" + "\n".join(
                r.get("_error", "?") for r in invalid[:5])
            if len(invalid) > 5:
                msg += f"\n... ve {len(invalid) - 5} satır daha"
        QMessageBox.warning(parent, "Aktarılacak Veri Yok", msg)
        return False

    # Özet + onay — mükerrer güncelleme seçeneği onay kutusunun içinde
    parts = [f"{len(valid)} yeni {label}"]
    if duplicates:
        parts.append(f"{len(duplicates)} mükerrer kayıt")
    if invalid:
        parts.append(f"{len(invalid)} hatalı satır (atlanacak)")
    text = Path(path).name + "\n\n" + "\n".join(f"• {p}" for p in parts)
    if invalid:
        text += "\n\nHatalı satır örnekleri:\n" + "\n".join(
            f"  - {r.get('_error', '?')}" for r in invalid[:3])
    text += "\n\nAktarım başlatılsın mı?"

    box = QMessageBox(parent)
    box.setWindowTitle("İçe Aktarma Onayı")
    box.setIcon(QMessageBox.Icon.Question)
    box.setText(text)
    chk = None
    if duplicates:
        chk = QCheckBox("Mükerrer kayıtları dosyadaki verilerle güncelle")
        chk.setToolTip("İşaretlenmezse mükerrer kayıtlar olduğu gibi bırakılır.")
        box.setCheckBox(chk)
    btn_ok = box.addButton("Aktar", QMessageBox.ButtonRole.AcceptRole)
    box.addButton("İptal", QMessageBox.ButtonRole.RejectRole)
    box.setDefaultButton(btn_ok)
    box.exec()
    if box.clickedButton() is not btn_ok:
        return False

    update_dups = bool(chk and chk.isChecked())
    rows = list(valid) + (list(duplicates) if update_dups else [])
    # ── AŞAMA: YAZMA — transaction ya tümüyle işler ya geri döner ────────
    prog = _ImportProgress(parent, f"{label.capitalize()} kaydediliyor…")
    asama = {}
    try:
        added, updated, skipped, errors = _perform_import(
            import_type, rows, update_dups, progress=prog, stage_state=asama)
    except Exception as exc:                                   # noqa: BLE001
        # Satır dışı fatal hata (ör. commit): ürün transaction'ı geri döndü.
        # Ama kategoriler transaction'dan ÖNCE yazılmış olabilir; tamamlanan
        # o aşama İNKÂR EDİLMEZ ve DB değiştiyse çağıran cache'i yeniler (18b).
        prog.close()
        kategori_yazildi = int(asama.get("kategori_yazildi", 0))
        if kategori_yazildi:
            # TEK kutu: hem tamamlanan aşamayı hem tamamlanamayanı söyler.
            hata_diyalogu.kismi_hata_goster(
                parent, "Kısmen Tamamlandı", exc,
                f"{kategori_yazildi} kategori oluşturuldu. "
                f"{label.capitalize()} aktarımı tamamlanamadı; "
                f"{label} kayıtları yazılmadı.",
                "Ice aktarma kaydet")
            return True
        hata_diyalogu.hata_goster(parent, "Hata", exc, "İçe aktarma", "kaydet")
        return False
    prog.close()

    kategori_yazildi = int(asama.get("kategori_yazildi", 0))
    parts = []
    if kategori_yazildi:
        # Kategoriler ayrı bir aşamadır; ürün sayısı 0 olsa bile DB değişti.
        parts.append(f"{kategori_yazildi} kategori oluşturuldu")
    if added:
        parts.append(f"{added} {label} eklendi")
    if updated:
        parts.append(f"{updated} {label} güncellendi")
    dup_skipped = len(duplicates) if not update_dups else 0
    if dup_skipped:
        parts.append(f"{dup_skipped} mükerrer atlandı")
    if skipped or invalid:
        parts.append(f"{skipped + len(invalid)} satır atlandı")
    if not parts:
        parts.append("İşlem yapılmadı.")
    msg = "\n".join(parts)
    if errors:
        msg += "\n\nHatalar:\n" + "\n".join(errors[:10])
        if len(errors) > 10:
            msg += f"\n... ve {len(errors) - 10} hata daha"
    QMessageBox.information(parent, "İçe Aktarma Tamamlandı", msg)
    logger.info("Excel import: type=%s added=%d updated=%d skipped=%d errors=%d",
                import_type, added, updated, skipped, len(errors))
    # Dönüş değeri YALNIZ "DB gerçekten değişti mi" sorusunu yanıtlar; çağıran
    # buna bakarak cache/ekran yeniler. Yalnız atlanan veya mükerrer satırlar
    # DB'yi değiştirmez.
    return bool(added or updated or kategori_yazildi)


def export_data_interactive(parent, import_type: str):
    """Kayıtlı verileri içe aktarmayla birebir uyumlu Excel'e yazar.

    Veri yoksa yalnızca başlık satırı yazılır — boş şablon işlevi görür.
    """
    import datetime
    today = datetime.date.today().strftime('%Y%m%d')
    try:
        if import_type == "customers":
            from services.customer_service import CustomerService
            records = CustomerService().get_all()
            default = f"musteriler_{today}.xlsx"
            label = "müşteri"
        elif import_type == "offers":
            from services.offer_service import OfferService
            svc = OfferService()
            # Kalemler dahil tam veri — roundtrip için get_by_id ile yüklenir
            records = [svc.get_by_id(o.id) for o in svc.get_all()]
            records = [o for o in records if o]
            default = f"teklifler_{today}.xlsx"
            label = "teklif"
        else:
            from services.product_service import ProductService
            records = ProductService().get_all()
            default = f"urunler_{today}.xlsx"
            label = "ürün"
    except Exception as exc:                                   # noqa: BLE001
        hata_diyalogu.hata_goster(parent, "Hata", exc, "Veri", "oku")
        return

    path, _ = QFileDialog.getSaveFileName(
        parent, "Veriyi Dışa Aktar", default, "Excel Dosyası (*.xlsx)")
    if not path:
        return
    try:
        if import_type == "customers":
            from services.export_service import export_customers_excel
            out = export_customers_excel(records, path)
        elif import_type == "offers":
            from services.export_service import export_offers_full_excel
            out = export_offers_full_excel(records, path)
        else:
            from services.export_service import export_products_excel
            from services.category_service import CategoryService
            cats = {c.id: c.name for c in CategoryService().get_all()}
            out = export_products_excel(records, path, cats)
        QMessageBox.information(
            parent, "Tamamlandı",
            f"{len(records)} {label} dışa aktarıldı.\n{out}\n\n"
            "Bu dosya 'Dosya → İçeri Aktar' ile olduğu gibi geri yüklenebilir.")
    except Exception as exc:                                   # noqa: BLE001
        # Veri OKUMA başarılıydı; yalnız dosyaya yazma başarısız oldu (18b).
        hata_diyalogu.hata_goster(parent, "Hata", exc, "Dışa aktarma", "kaydet")


# ── Teklif içe aktarma (kalemli format) ───────────────────────────────────────
# export_offers_full_excel çıktısıyla birebir uyumlu: her satır bir kalem,
# satırlar "Teklif No"ya göre gruplanıp teklif yeniden kurulur.

OFFER_MAP = {
    "teklif no": "offer_no",
    "firma": "company_name",         "firma adı": "company_name",
    "ilgili kişi": "contact_person",
    "adres": "address",
    "telefon": "phone",
    "e-posta": "email",              "eposta": "email",
    "tarih": "date",
    "para birimi": "currency",
    "durum": "status",
    "geçerlilik": "validity",
    "geçerlilik notu": "validity_note",
    "ödeme": "payment_term",         "ödeme vadesi": "payment_term",
    "iskonto (%)": "discount_percent", "iskonto": "discount_percent",
    "ürün kodu": "product_code",     "kod": "product_code",
    "ürün adı": "product_name",      "ürün": "product_name",
    "ürün açıklama": "item_description", "açıklama": "item_description",
    "adet": "quantity",              "miktar": "quantity",
    "birim": "unit",
    "teslim süresi": "delivery_time", "teslim": "delivery_time",
    "birim fiyat": "unit_price",     "fiyat": "unit_price",
}


def _validate_offer_rows(raw_rows: list) -> tuple[list, list, list]:
    """Satırları teklife gruplar. (yeni_teklifler, mükerrer_nolar, hatalar)"""
    from database.db_manager import get_db
    from core.constants import STATUS_ORDER
    from core.offer_files import validate_offer_number
    db = get_db()

    groups, order, invalid, invalid_offer_numbers = {}, [], [], set()
    for idx, raw in enumerate(raw_rows, 2):   # 1. satır başlık
        r = _map_row(raw, OFFER_MAP)
        no = str(r.get("offer_no", "")).strip()
        if not no:
            invalid.append(f"Satır {idx}: Teklif No eksik")
            continue
        try:
            no = validate_offer_number(no)
        except ValueError:
            invalid.append(f"Satır {idx}: Teklif No geçersiz")
            continue
        g = groups.get(no)
        if g is None:
            company = str(r.get("company_name", "")).strip()
            if not company:
                invalid.append(f"Satır {idx}: Firma adı eksik ({no})")
                continue
            status = str(r.get("status", "")).strip()
            g = {
                "offer_no": no,
                "company_name": company,
                "contact_person": str(r.get("contact_person", "")).strip(),
                "address": str(r.get("address", "")).strip(),
                "phone": str(r.get("phone", "")).strip(),
                "email": str(r.get("email", "")).strip(),
                "date": str(r.get("date", "")).strip(),
                "currency": normalize_currency(r.get("currency", ""),
                                               default="EUR"),
                "status": status if status in STATUS_ORDER else "Beklemede",
                "validity": str(r.get("validity", "")).strip(),
                "validity_note": str(r.get("validity_note", "")).strip(),
                "payment_term": str(r.get("payment_term", "")).strip(),
                "discount_percent": _parse_number(r.get("discount_percent"), 0),
                "items": [],
            }
            groups[no] = g
            order.append(no)
        pname = str(r.get("product_name", "")).strip()
        pcode = str(r.get("product_code", "")).strip()
        if not (pname or pcode):
            invalid.append(f"Satır {idx}: Ürün bilgisi eksik ({no})")
            continue
        quantity_value = r.get("quantity", "")
        if quantity_value is None or (isinstance(quantity_value, str)
                                      and not quantity_value.strip()):
            qty = 1
        else:
            qty = _parse_number(quantity_value, None)
            if qty is None or qty <= 0:
                invalid.append(f"Satır {idx}: Miktar sıfırdan büyük sayı olmalıdır")
                invalid_offer_numbers.add(no)
                continue
        price = _parse_number(r.get("unit_price"), 0)
        g["items"].append({
            "product_code": pcode,
            "product_name": pname or pcode,
            "description": str(r.get("item_description", "")).strip(),
            "quantity": qty,
            "unit": str(r.get("unit", "")).strip() or "Adet",
            "delivery_time": str(r.get("delivery_time", "")).strip() or "2-3 Hafta",
            "unit_price": price,
        })

    new_offers, dups = [], []
    for no in order:
        if db.fetchone("SELECT id FROM offers WHERE offer_no = ?", (no,)):
            dups.append(no)
        elif no in invalid_offer_numbers:
            continue
        elif not groups[no]["items"]:
            invalid.append(f"{no}: hiç geçerli ürün kalemi yok")
        else:
            new_offers.append(groups[no])
    return new_offers, dups, invalid


def _perform_offer_import(offer_groups: list) -> tuple[int, list]:
    """Gruplanmış teklifleri kaydeder. (eklenen, hatalar) döner."""
    import datetime as _dt
    from services.offer_service import OfferService
    from models.offer import Offer, calculate_discount
    from models.offer_item import OfferItem

    svc = OfferService()
    added, errors = 0, []
    for sira, g in enumerate(offer_groups, 1):
        try:
            items = [OfferItem(
                product_code=i["product_code"], product_name=i["product_name"],
                description=i["description"], quantity=i["quantity"],
                unit=i["unit"], delivery_time=i["delivery_time"],
                unit_price=i["unit_price"],
                total_price=i["quantity"] * i["unit_price"],
            ) for i in g["items"]]
            subtotal = sum(it.total_price for it in items)
            pct = float(g["discount_percent"] or 0)
            discount = calculate_discount(subtotal, "percent", pct)
            offer = Offer(
                offer_no=g["offer_no"], company_name=g["company_name"],
                contact_person=g["contact_person"],
                customer_address=g["address"], customer_phone=g["phone"],
                customer_email=g["email"],
                date=g["date"] or _dt.date.today().strftime("%d.%m.%Y"),
                currency=g["currency"], status=g["status"],
                validity=g["validity"], validity_note=g["validity_note"],
                payment_term=g["payment_term"],
                discount_type="percent", discount_value=pct,
                discount_amount=discount, show_discount=pct > 0,
                total_amount=subtotal - discount, items=items,
            )
            svc.save(offer, keep_offer_no=True)
            added += 1
        except Exception as exc:                               # noqa: BLE001
            # Teklif no, firma verisi ve ham istisna kullanıcıya GİTMEZ;
            # güvenli kayit_id olarak yalnız grup SIRASI kullanılır. Bir
            # grubun hatası sonraki grupları engellemez.
            op_hata.logla(exc, "Teklif yaz", kayit_id=sira)
            errors.append(TEKLIF_YAZILAMADI.format(sira=sira))
    return added, errors


def _run_offer_import_flow(parent, path: str, raw_rows: list) -> bool:
    """Teklif dosyası için onay + aktarım."""
    new_offers, dups, invalid = _validate_offer_rows(raw_rows)
    if not new_offers:
        msg = "Dosyada aktarılabilir yeni teklif bulunamadı."
        if dups:
            ekstra = ", ".join(dups[:5]) + (" ..." if len(dups) > 5 else "")
            msg += f"\n\n{len(dups)} teklif zaten kayıtlı (atlandı):\n{ekstra}"
        if invalid:
            msg += "\n\nHatalı satırlar:\n" + "\n".join(invalid[:5])
        QMessageBox.warning(parent, "Aktarılacak Veri Yok", msg)
        return False

    item_count = sum(len(g["items"]) for g in new_offers)
    parts = [f"{len(new_offers)} yeni teklif ({item_count} kalem)"]
    if dups:
        parts.append(f"{len(dups)} teklif zaten kayıtlı (atlanacak)")
    if invalid:
        parts.append(f"{len(invalid)} hatalı satır (atlanacak)")
    text = Path(path).name + "\n\n" + "\n".join(f"• {p}" for p in parts)
    if invalid:
        text += "\n\nHatalı satır örnekleri:\n" + "\n".join(
            f"  - {dogrulama}" for dogrulama in invalid[:3])
    text += "\n\nAktarım başlatılsın mı?"

    box = QMessageBox(parent)
    box.setWindowTitle("Teklif İçe Aktarma Onayı")
    box.setIcon(QMessageBox.Icon.Question)
    box.setText(text)
    btn_ok = box.addButton("Aktar", QMessageBox.ButtonRole.AcceptRole)
    box.addButton("İptal", QMessageBox.ButtonRole.RejectRole)
    box.setDefaultButton(btn_ok)
    box.exec()
    if box.clickedButton() is not btn_ok:
        return False

    added, errors = _perform_offer_import(new_offers)
    msg = f"{added} teklif eklendi"
    if dups:
        msg += f"\n{len(dups)} mevcut teklif atlandı"
    if invalid:
        msg += f"\n{len(invalid)} hatalı satır atlandı"
    if errors:
        msg += "\n\nHatalar:\n" + "\n".join(errors[:10])
    QMessageBox.information(parent, "İçe Aktarma Tamamlandı", msg)
    logger.info("Teklif import: added=%d dup=%d invalid=%d errors=%d",
                added, len(dups), len(invalid), len(errors))
    return added > 0


# ── Tümünü içe / dışa aktar (tek dosya, 3 sayfa) ─────────────────────────────

def _read_xlsx_sheets(path: str) -> tuple[dict, str]:
    """Çok sayfalı Excel okur: ({sayfa_adı: satırlar}, hata)."""
    try:
        import openpyxl
    except ImportError:
        return {}, "openpyxl kütüphanesi bulunamadı."
    # Kapatma GERÇEK `finally` içindedir: okuma yarıda hata verirse de
    # workbook açık kalmaz (Windows'ta dosya kilidi bırakırdı).
    wb = None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        result = {}
        for ws in wb.worksheets:
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                continue
            headers = [str(c or "").strip() for c in all_rows[0]]
            rows = []
            for row in all_rows[1:]:
                if all(c is None for c in row):
                    continue
                rows.append({headers[i]: (str(v) if v is not None else "")
                             for i, v in enumerate(row) if i < len(headers)})
            result[ws.title] = rows
        return result, ""
    except Exception as exc:                                   # noqa: BLE001
        op_hata.logla(exc, "Dosya oku")
        return {}, DOSYA_OKUMA_HATASI
    finally:
        # Kapatma hatası ne başarıyı ne de asıl okuma hatasını değiştirir.
        _workbook_kapat(wb)


def run_import_all_flow(parent) -> bool:
    """Tek dosyadan Müşteri + Ürün + Teklif verisini birlikte içe aktarır.

    Dosyadaki 'Müşteriler', 'Ürünler', 'Teklifler' sayfaları tanınır;
    eksik sayfalar atlanır. Sıra: müşteriler → ürünler → teklifler.
    """
    path, _ = QFileDialog.getOpenFileName(
        parent, "Tümünü İçe Aktar (Tek Dosya)", "",
        "Excel Dosyası (*.xlsx *.xlsm)")
    if not path:
        return False

    sheets, err = _read_xlsx_sheets(path)
    if err:
        QMessageBox.warning(parent, "Dosya Hatası", err)
        return False

    def _find(keyword):
        for name, rows in sheets.items():
            if keyword in _norm(name):
                return rows
        return None

    cust_rows = _find("müşteri")
    prod_rows = _find("ürün")
    off_rows  = _find("teklif")
    if cust_rows is None and prod_rows is None and off_rows is None:
        QMessageBox.warning(
            parent, "Sayfa Bulunamadı",
            "Bu dosyada 'Müşteriler', 'Ürünler' veya 'Teklifler' adlı sayfa "
            "bulunamadı.\n\nTek türlü bir dosya aktarmak istiyorsanız "
            "Dosya → İçeri Aktar menüsündeki ilgili seçeneği kullanın.")
        return False

    # ── Doğrulama ────────────────────────────────────────────────────────
    prog = _ImportProgress(parent, "Kayıtlar denetleniyor…")
    c_valid = c_dup = p_valid = p_dup = []
    o_new, o_dup, all_invalid = [], [], []
    try:
        if cust_rows:
            prog.set_label("Müşteriler denetleniyor…")
            c_valid, c_dup, c_inv = _validate_rows("customers", cust_rows,
                                                   progress=prog)
            all_invalid += [f"Müşteri: {r.get('_error','?')}" for r in c_inv]
        if prod_rows:
            prog.set_label("Ürünler denetleniyor…")
            p_valid, p_dup, p_inv = _validate_rows("products", prod_rows,
                                                   progress=prog)
            all_invalid += [f"Ürün: {r.get('_error','?')}" for r in p_inv]
        if off_rows:
            o_new, o_dup, o_inv = _validate_offer_rows(off_rows)
            all_invalid += [f"Teklif: {d}" for d in o_inv]
    except Exception as exc:                                   # noqa: BLE001
        # Doğrulama aşaması: DB'ye HENÜZ hiçbir şey yazılmadı.
        prog.close()
        hata_diyalogu.hata_goster(parent, "Hata", exc, "İçe aktarma", "denetle")
        return False
    prog.close()

    if not (c_valid or c_dup or p_valid or p_dup or o_new):
        QMessageBox.warning(parent, "Aktarılacak Veri Yok",
                            "Dosyada aktarılabilir yeni kayıt bulunamadı.")
        return False

    # ── Özet + onay ──────────────────────────────────────────────────────
    parts = []
    if cust_rows is not None:
        parts.append(f"{len(c_valid)} yeni müşteri"
                     + (f", {len(c_dup)} mükerrer" if c_dup else ""))
    if prod_rows is not None:
        parts.append(f"{len(p_valid)} yeni ürün"
                     + (f", {len(p_dup)} mükerrer" if p_dup else ""))
    if off_rows is not None:
        item_count = sum(len(g["items"]) for g in o_new)
        parts.append(f"{len(o_new)} yeni teklif ({item_count} kalem)"
                     + (f", {len(o_dup)} zaten kayıtlı" if o_dup else ""))
    if all_invalid:
        parts.append(f"{len(all_invalid)} hatalı satır (atlanacak)")
    text = Path(path).name + "\n\n" + "\n".join(f"• {p}" for p in parts)
    text += "\n\nAktarım başlatılsın mı?"

    box = QMessageBox(parent)
    box.setWindowTitle("Tümünü İçe Aktarma Onayı")
    box.setIcon(QMessageBox.Icon.Question)
    box.setText(text)
    chk = None
    if c_dup or p_dup:
        chk = QCheckBox("Mükerrer müşteri/ürün kayıtlarını dosyayla güncelle")
        box.setCheckBox(chk)
    btn_ok = box.addButton("Aktar", QMessageBox.ButtonRole.AcceptRole)
    box.addButton("İptal", QMessageBox.ButtonRole.RejectRole)
    box.setDefaultButton(btn_ok)
    box.exec()
    if box.clickedButton() is not btn_ok:
        return False

    update_dups = bool(chk and chk.isChecked())

    # ── Aktarım: müşteri → ürün → teklif ─────────────────────────────────
    # Müşteri / ürün / teklif AYRI aşamalardır ve ayrı transaction kullanır.
    # Bir aşamanın düşmesi diğerlerinin verisini bozmaz; bu yüzden sonraki
    # aşama ATLANMAZ — kullanıcının aktarmayı istediği veri sessizce
    # düşürülmez. Tamamlanan aşama "başarısız" gibi anlatılmaz (18b).
    prog = _ImportProgress(parent, "Kaydediliyor…")
    summary, errors = [], []
    yazildi = False
    try:
        if cust_rows:
            prog.set_label("Müşteriler kaydediliyor…")
            rows = list(c_valid) + (list(c_dup) if update_dups else [])
            try:
                a, u, s, e = _perform_import("customers", rows, update_dups,
                                             progress=prog)
                summary.append(f"Müşteri: {a} eklendi"
                               + (f", {u} güncellendi" if u else ""))
                errors += e
                yazildi = yazildi or bool(a or u)
            except Exception as exc:                           # noqa: BLE001
                op_hata.logla(exc, "Musteri asamasi")
                summary.append(ASAMA_BASARISIZ["customers"])
        if prod_rows:
            prog.set_label("Ürünler kaydediliyor…")
            rows = list(p_valid) + (list(p_dup) if update_dups else [])
            asama = {}
            try:
                a, u, s, e = _perform_import("products", rows, update_dups,
                                             progress=prog, stage_state=asama)
                urun_ozet = (f"Ürün: {a} eklendi"
                             + (f", {u} güncellendi" if u else ""))
                errors += e
                yazildi = yazildi or bool(a or u)
            except Exception as exc:                           # noqa: BLE001
                op_hata.logla(exc, "Urun asamasi")
                urun_ozet = ASAMA_BASARISIZ["products"]
            # Kategoriler ürün transaction'ından BAĞIMSIZ yazılır: transaction
            # düşse bile DB değişmiş olabilir → özette görünür, yenileme gerekir.
            kategori_yazildi = int(asama.get("kategori_yazildi", 0))
            if kategori_yazildi:
                summary.append(f"Kategori: {kategori_yazildi} oluşturuldu")
                yazildi = True
            summary.append(urun_ozet)
        if off_rows:
            prog.set_label("Teklifler kaydediliyor…")
            try:
                a, e = _perform_offer_import(o_new)
                summary.append(f"Teklif: {a} eklendi"
                               + (f", {len(o_dup)} mevcut atlandı" if o_dup else ""))
                errors += e
                yazildi = yazildi or bool(a)
            except Exception as exc:                           # noqa: BLE001
                op_hata.logla(exc, "Teklif asamasi")
                summary.append(ASAMA_BASARISIZ["offers"])
    finally:
        prog.close()

    msg = "\n".join(summary)
    if all_invalid:
        msg += f"\n{len(all_invalid)} hatalı satır atlandı"
    if errors:
        msg += "\n\nHatalar:\n" + "\n".join(errors[:10])
    QMessageBox.information(parent, "İçe Aktarma Tamamlandı", msg)
    logger.info("Tumunu import: %s | hatalar=%d", " / ".join(summary), len(errors))
    # DB gerçekten değiştiyse çağıran ekran/cache yenilemesi yapar.
    return yazildi


def export_all_interactive(parent):
    """Müşteri + Ürün + Teklif verisini tek dosyada dışa aktarır."""
    import datetime
    try:
        from services.customer_service import CustomerService
        from services.product_service import ProductService
        from services.offer_service import OfferService
        from services.category_service import CategoryService
        customers = CustomerService().get_all()
        products = ProductService().get_all()
        svc = OfferService()
        offers_full = [o for o in (svc.get_by_id(x.id) for x in svc.get_all()) if o]
        cats = {c.id: c.name for c in CategoryService().get_all()}
    except Exception as exc:                                   # noqa: BLE001
        hata_diyalogu.hata_goster(parent, "Hata", exc, "Veri", "oku")
        return

    default = f"teklif_yonetim_verileri_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
    path, _ = QFileDialog.getSaveFileName(
        parent, "Tümünü Dışa Aktar (Tek Dosya)", default,
        "Excel Dosyası (*.xlsx)")
    if not path:
        return
    try:
        from services.export_service import export_all_excel
        out = export_all_excel(path, customers, products, offers_full, cats)
        QMessageBox.information(
            parent, "Tamamlandı",
            f"{len(customers)} müşteri, {len(products)} ürün, "
            f"{len(offers_full)} teklif tek dosyaya aktarıldı.\n{out}\n\n"
            "Bu dosya 'Dosya → İçeri Aktar → Tümünü İçe Aktar' ile "
            "olduğu gibi geri yüklenebilir.")
    except Exception as exc:                                   # noqa: BLE001
        hata_diyalogu.hata_goster(parent, "Hata", exc, "Dışa aktarma", "kaydet")
