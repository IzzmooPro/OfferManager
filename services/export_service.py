"""Excel ve CSV export servisi."""
import csv, logging
from pathlib import Path
from typing import List
from core.constants import SYM_MAP
from core.formatting import fmt_money
from core.date_utils import to_display_date

logger = logging.getLogger("export_service")

HEADERS = ["Teklif No", "Firma", "Tarih", "Para Birimi", "Toplam Tutar",
           "Durum", "Vade", "Ödeme", "İlgili Kişi"]

def _value(record, key: str, default=""):
    if isinstance(record, dict):
        value = record.get(key, default)
    else:
        value = getattr(record, key, default)
    return default if value is None else value


def _row(o) -> list:
    currency = _value(o, "currency", "")
    sym = SYM_MAP.get(currency, "")
    return [
        _value(o, "offer_no", ""),
        _value(o, "company_name", ""),
        to_display_date(_value(o, "date", "")),
        currency,
        fmt_money(_value(o, 'total_amount', 0), sym),
        _value(o, "status", ""),
        _value(o, "validity", ""),
        _value(o, "payment_term", ""),
        _value(o, "contact_person", ""),
    ]


def _currency_totals(offers) -> str:
    totals = {}
    for offer in offers:
        currency = _value(offer, "currency", "") or "?"
        totals[currency] = totals.get(currency, 0.0) + float(
            _value(offer, "total_amount", 0) or 0)
    parts = []
    for currency in ("TL", "EUR", "USD"):
        if currency in totals:
            parts.append(fmt_money(totals.pop(currency), SYM_MAP.get(currency, currency)))
    parts.extend(fmt_money(amount, currency) for currency, amount in sorted(totals.items()))
    return " | ".join(parts)


def export_excel(offers: List[object], path: str) -> str:
    """Verilen teklif listesini Excel'e yazar. path döner."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl kurulu değil: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Teklifler"

    # Başlık satırı
    header_fill = PatternFill("solid", fgColor="1E4D8C")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.row_dimensions[1].height = 20

    # Veri satırları
    alt_fill = PatternFill("solid", fgColor="F0F4FA")
    for r_idx, o in enumerate(offers, 2):
        fill = alt_fill if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for c_idx, val in enumerate(_row(o), 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill   = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[r_idx].height = 16

    # Sütun genişlikleri
    widths = [18, 28, 12, 10, 16, 12, 10, 14, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Özet satırı
    sum_row = len(offers) + 2
    ws.cell(row=sum_row, column=1, value="TOPLAM").font = Font(bold=True)
    ws.cell(row=sum_row, column=5,
            value=_currency_totals(offers)).font = Font(bold=True)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    logger.info("Excel export: %s (%d teklif)", path, len(offers))
    return path


def export_csv(offers: List[object], path: str) -> str:
    """Verilen teklif listesini CSV'ye yazar. path döner."""
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(HEADERS)
        for o in offers:
            w.writerow(_row(o))
    logger.info("CSV export: %s (%d teklif)", path, len(offers))
    return path


# ── Ürün / Müşteri dışa aktarma ───────────────────────────────────────────────
# Başlıklar excel_import.py'deki eşleştirme haritalarıyla BİREBİR uyumludur;
# dışa aktarılan dosya değiştirilmeden yeniden içe aktarılabilir (roundtrip).
# Fiyat/stok sayısal hücre olarak yazılır — biçimlendirilmiş metin değil.

PRODUCT_EXPORT_HEADERS = ["Ürün Kodu", "Ürün Adı", "Açıklama", "Fiyat",
                          "Para Birimi", "Stok", "Birim", "Kategori"]
CUSTOMER_EXPORT_HEADERS = ["Firma Adı", "İlgili Kişi", "Adres", "Telefon",
                           "E-posta", "Not"]


def _write_table_sheet(ws, headers: list, rows: list) -> None:
    """Verilen çalışma sayfasına stilli başlık + veri satırları yazar."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1E4D8C")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 20

    alt_fill = PatternFill("solid", fgColor="F0F4FA")
    for r_idx, row in enumerate(rows, 2):
        fill = alt_fill if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # Sütun genişliği içeriğe göre — hiçbir hücre görsel olarak kesilmesin
    for i, header in enumerate(headers, 1):
        longest = len(str(header))
        for row in rows:
            val = row[i - 1]
            text = f"{val:.2f}" if isinstance(val, float) else str(val)
            longest = max(longest, len(text))
        ws.column_dimensions[get_column_letter(i)].width = min(70, longest + 4)

    # Başlık satırı sabit kalsın + filtre okları
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"


def _export_table_excel(path: str, sheet_title: str, headers: list,
                        rows: list) -> str:
    """Tek sayfalık stilli Excel dosyası yazar."""
    try:
        from openpyxl import Workbook
    except ImportError:
        raise RuntimeError("openpyxl kurulu değil: pip install openpyxl")
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    _write_table_sheet(ws, headers, rows)
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _product_rows(products: List[object], category_names: dict) -> list:
    return [[
        p.product_code or "",
        p.product_name or "",
        p.description or "",
        float(p.price or 0),
        p.currency or "",
        float(p.stock or 0),
        p.unit or "",
        category_names.get(p.category_id, ""),
    ] for p in products]


def export_products_excel(products: List[object], path: str,
                          category_names: dict | None = None) -> str:
    """Ürün kataloğunu içe aktarmayla uyumlu Excel'e yazar."""
    rows = _product_rows(products, category_names or {})
    out = _export_table_excel(path, "Ürünler", PRODUCT_EXPORT_HEADERS, rows)
    logger.info("Ürün export: %s (%d ürün)", out, len(rows))
    return out


def _customer_rows(customers: List[object]) -> list:
    return [[
        c.company_name or "",
        c.contact_person or "",
        c.address or "",
        c.phone or "",
        c.email or "",
        getattr(c, "notes", "") or "",
    ] for c in customers]


def export_customers_excel(customers: List[object], path: str) -> str:
    """Müşteri listesini içe aktarmayla uyumlu Excel'e yazar."""
    rows = _customer_rows(customers)
    out = _export_table_excel(path, "Müşteriler", CUSTOMER_EXPORT_HEADERS, rows)
    logger.info("Müşteri export: %s (%d müşteri)", out, len(rows))
    return out


# ── Teklif tam dışa aktarma (kalemli — içe aktarmayla uyumlu) ────────────────
# Her kalem bir satırdır; teklif alanları o teklifin tüm satırlarında tekrar
# eder. İçe aktarma satırları "Teklif No"ya göre gruplayıp teklifi yeniden
# kurar — dosya değiştirilmeden geri yüklenebilir.

OFFER_FULL_EXPORT_HEADERS = [
    "Teklif No", "Firma", "İlgili Kişi", "Adres", "Telefon", "E-posta",
    "Tarih", "Para Birimi", "Durum", "Geçerlilik", "Geçerlilik Notu",
    "Ödeme", "İskonto (%)",
    "Ürün Kodu", "Ürün Adı", "Ürün Açıklama", "Adet", "Birim",
    "Teslim Süresi", "Birim Fiyat",
]


def _offer_rows(offers: List[object]) -> list:
    rows = []
    for o in offers:
        pct = float(o.discount_value or 0) if (o.discount_type or "") == "percent" else 0.0
        for item in (o.items or []):
            rows.append([
                o.offer_no or "",
                o.company_name or "",
                o.contact_person or "",
                o.customer_address or "",
                o.customer_phone or "",
                o.customer_email or "",
                to_display_date(o.date),
                o.currency or "",
                o.status or "Beklemede",
                o.validity or "",
                o.validity_note or "",
                o.payment_term or "",
                pct,
                item.product_code or "",
                item.product_name or "",
                item.description or "",
                float(item.quantity or 0),
                item.unit or "",
                item.delivery_time or "",
                float(item.unit_price or 0),
            ])
    return rows


def export_offers_full_excel(offers: List[object], path: str) -> str:
    """Teklifleri kalemleriyle birlikte içe aktarmayla uyumlu Excel'e yazar.

    offers: items alanı YÜKLÜ Offer nesneleri (get_by_id ile alınmış).
    """
    rows = _offer_rows(offers)
    out = _export_table_excel(path, "Teklifler", OFFER_FULL_EXPORT_HEADERS, rows)
    logger.info("Teklif tam export: %s (%d teklif, %d kalem)",
                out, len(offers), len(rows))
    return out


def export_all_excel(path: str, customers: List[object],
                     products: List[object], offers_full: List[object],
                     category_names: dict | None = None) -> str:
    """Müşteri + Ürün + Teklif verisini TEK dosyada 3 sayfa olarak yazar.

    Her sayfa, tekli içe aktarmayla birebir aynı formattadır; dosya
    'Tümünü İçe Aktar' ile olduğu gibi geri yüklenebilir.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        raise RuntimeError("openpyxl kurulu değil: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Müşteriler"
    _write_table_sheet(ws, CUSTOMER_EXPORT_HEADERS, _customer_rows(customers))
    ws = wb.create_sheet("Ürünler")
    _write_table_sheet(ws, PRODUCT_EXPORT_HEADERS,
                       _product_rows(products, category_names or {}))
    ws = wb.create_sheet("Teklifler")
    _write_table_sheet(ws, OFFER_FULL_EXPORT_HEADERS, _offer_rows(offers_full))

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    logger.info("Tümü export: %s (%d müşteri, %d ürün, %d teklif)",
                path, len(customers), len(products), len(offers_full))
    return path
