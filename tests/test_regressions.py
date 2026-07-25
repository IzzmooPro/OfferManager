"""Kritik finansal veri, export, yedek ve import regresyon testleri."""
import ast
import csv
import inspect
import os
import re
import sqlite3
import tempfile
import unittest
import zipfile
from datetime import date
from pathlib import Path


_DATA_ROOT = tempfile.TemporaryDirectory(prefix="oms_regression_data_")
os.environ["LOCALAPPDATA"] = _DATA_ROOT.name

from core.app_paths import (
    DB_PATH, LOGO_DISABLED_PATH, LOGO_PATH,
    SIG1_PATH, SIG2_PATH, SIG3_PATH, SIG4_PATH,
)
from database.db_manager import get_db
from models.offer import Offer
from models.offer_item import OfferItem
from pdf.pdf_generator import (
    generate_pdf,
    _is_redundant_validity_note,
    _load_company,
)
from ui.create_offer_page import (
    _normalize_payment_text,
    _normalize_validity_text,
    _payment_days_for_input,
    _validity_days_for_input,
)
from services.export_service import export_csv, export_excel
from services.offer_service import OfferService
from ui.dialogs import backup_manager
from ui.utils.excel_import import _parse_number
from ui.utils.updater import is_newer_version


def _offer(currency="TL", amount=90.0):
    return Offer(
        company_name="Test Firma",
        date=date.today().strftime("%d.%m.%Y"),
        currency=currency,
        total_amount=amount,
        discount_amount=10.0,
        discount_type="percent",
        discount_value=10.0,
        items=[OfferItem(
            product_name="Test Ürün",
            quantity=1,
            unit_price=100.0,
            total_price=100.0,
        )],
    )


class RegressionTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.db = get_db()
        cls.service = OfferService()

    @classmethod
    def tearDownClass(cls):
        cls.db.close()
        _DATA_ROOT.cleanup()

    def setUp(self):
        with self.db.transaction() as conn:
            conn.execute("DELETE FROM offer_items")
            conn.execute("DELETE FROM offers")
            conn.execute("DELETE FROM offer_counter")

    def test_date_revenue_and_range(self):
        offer_id = self.service.save(_offer())
        stored = self.service.get_by_id(offer_id)
        self.assertEqual(stored.date, date.today().isoformat())
        self.assertEqual(stored.discount_amount, 10.0)
        self.assertEqual(stored.total_amount, 90.0)
        self.assertEqual(self.service.get_revenue_summary()["monthly"]["TL"], 90.0)
        self.assertEqual(len(self.service.get_by_date_range(
            date.today().replace(day=1).isoformat(), date.today().isoformat())), 1)

    def test_invalid_offer_is_rejected(self):
        with self.assertRaises(ValueError):
            self.service.save(Offer(company_name="Boş", date=date.today().isoformat()))
        invalid = _offer()
        invalid.discount_amount = 101
        invalid.total_amount = -1
        with self.assertRaises(ValueError):
            self.service.save(invalid)

        invalid_rate = _offer()
        invalid_rate.discount_type = "percent"
        invalid_rate.discount_value = 101
        with self.assertRaises(ValueError):
            self.service.save(invalid_rate)

    def test_percentage_discount_is_calculated_and_persisted(self):
        offer = _offer(amount=80)
        offer.discount_type = "percent"
        offer.discount_value = 20
        offer.discount_amount = 0
        offer.show_discount = False

        saved = self.service.get_by_id(self.service.save(offer))
        self.assertEqual(saved.discount_type, "percent")
        self.assertEqual(saved.discount_value, 20)
        self.assertEqual(saved.discount_amount, 20)
        self.assertEqual(saved.total_amount, 80)
        self.assertFalse(saved.show_discount)

    def test_pdf_requires_manual_validity_and_payment_terms(self):
        with tempfile.TemporaryDirectory(prefix="oms_pdf_validation_") as temp_dir:
            output = Path(temp_dir) / "offer.pdf"
            with self.assertRaisesRegex(ValueError, "Geçerlilik"):
                generate_pdf(Offer(payment_term="Peşin"), str(output))
            with self.assertRaisesRegex(ValueError, "Ödeme"):
                generate_pdf(Offer(validity="15 Gün"), str(output))
            self.assertFalse(output.exists())

    def test_pdf_compacts_small_overflow_to_one_page(self):
        offer = _offer()
        offer.validity = "10 Gün"
        offer.payment_term = "15 Gün Vadeli"
        offer.company_name = "Uzun Firma Adı Teknik Malzemeler Limited Şirketi"
        offer.customer_address = "Uzun Cadde No:63 Ümraniye İstanbul"
        offer.items = [
            OfferItem(
                product_code="PDFTEST-10-A",
                product_name="Endüstriyel Test Ürünü 10-A",
                description=(
                    "Paslanmaz gövdeli, yüksek performanslı test ürünü; model 10A."),
                quantity=3, unit="Kutu", delivery_time="Stokta Var",
                unit_price=100, total_price=300,
            ),
            OfferItem(
                product_code="PDFTEST-10-B",
                product_name="Yardımcı Ekipman Seti 10-B",
                description=(
                    "Montaj aksesuarları, bağlantı elemanları ve kullanım "
                    "dokümanları dahil komple set."),
                quantity=3, unit="Set", delivery_time="Stokta Var",
                unit_price=100, total_price=300,
            ),
        ]
        offer.total_amount = 600
        offer.discount_amount = 0

        with tempfile.TemporaryDirectory(prefix="oms_pdf_layout_") as temp_dir:
            output = Path(temp_dir) / "adaptive.pdf"
            generate_pdf(offer, str(output))
            page_count = len(re.findall(rb"/Type\s*/Page\b", output.read_bytes()))
            self.assertEqual(page_count, 1)

    def test_numeric_offer_terms_are_completed_for_pdf(self):
        self.assertEqual(_normalize_validity_text("10"), "10 Gün")
        self.assertEqual(_validity_days_for_input("10 Gün"), "10")
        self.assertEqual(_validity_days_for_input("31.12.2026"), "")
        self.assertEqual(_normalize_payment_text("45"), "45 Gün Vadeli")
        self.assertEqual(_normalize_payment_text("45 gün"), "45 Gün Vadeli")
        self.assertEqual(_payment_days_for_input("45 Gün Vadeli"), "45")
        self.assertEqual(_payment_days_for_input("Peşin"), "")
        self.assertTrue(_is_redundant_validity_note(
            "Fiyatlar teklif geçerlilik süresi boyunca sabittir."))
        self.assertFalse(_is_redundant_validity_note(
            "Hammadde fiyatlarındaki değişiklikler hariçtir."))

    def test_default_pdf_language_is_clear_turkish(self):
        company = _load_company()
        combined = " ".join(
            value for key, value in company.items() if key.startswith("pdf_"))
        for outdated in ("tetkik", "müteakip", "maille", "fax", "şartiyla"):
            self.assertNotIn(outdated, combined.casefold())
        self.assertIn("Türk lirasına", combined)

    def test_export_accepts_models_and_separates_currencies(self):
        first = self.service.get_by_id(self.service.save(_offer("TL", 90)))
        second_offer = _offer("EUR", 190)
        second_offer.items[0].total_price = 200
        second_offer.items[0].unit_price = 200
        second_offer.total_amount = 180
        second = self.service.get_by_id(self.service.save(second_offer))
        with tempfile.TemporaryDirectory(prefix="oms_export_") as temp_dir:
            xlsx = Path(temp_dir) / "offers.xlsx"
            csv_path = Path(temp_dir) / "offers.csv"
            export_excel([first, second], str(xlsx))
            export_csv([first, second], str(csv_path))
            from openpyxl import load_workbook
            summary = load_workbook(xlsx, data_only=True).active.cell(4, 5).value
            self.assertIn("|", summary)
            with csv_path.open(encoding="utf-8-sig") as stream:
                self.assertEqual(len(list(csv.reader(stream, delimiter=";"))), 3)

    def test_backup_is_unique_consistent_and_validated(self):
        self.service.save(_offer())
        with tempfile.TemporaryDirectory(prefix="oms_backup_") as temp_dir:
            first = Path(backup_manager.create_backup(temp_dir))
            second = Path(backup_manager.create_backup(temp_dir))
            self.assertNotEqual(first, second)

            connection = sqlite3.connect(DB_PATH)
            connection.execute("PRAGMA journal_mode=WAL")
            connection.execute("INSERT INTO customers(company_name) VALUES ('WAL')")
            connection.commit()
            wal_backup = Path(backup_manager.create_backup(temp_dir))
            connection.close()
            with zipfile.ZipFile(wal_backup) as archive:
                snapshot = Path(temp_dir) / "snapshot.db"
                snapshot.write_bytes(archive.read("database.db"))
            copied = sqlite3.connect(snapshot)
            count = copied.execute(
                "SELECT COUNT(*) FROM customers WHERE company_name='WAL'").fetchone()[0]
            copied.close()
            self.assertEqual(count, 1)

            corrupt = Path(temp_dir) / "corrupt.zip"
            with zipfile.ZipFile(corrupt, "w") as archive:
                archive.writestr("database.db", b"not sqlite")
            with self.assertRaises(ValueError):
                backup_manager.restore_backup(str(corrupt))

    def test_restore_recreates_exact_asset_state(self):
        self.service.save(_offer())
        for path in (LOGO_PATH, LOGO_DISABLED_PATH):
            if path.exists():
                path.unlink()
        with tempfile.TemporaryDirectory(prefix="oms_assets_") as temp_dir:
            clean = backup_manager.create_backup(temp_dir)
            LOGO_PATH.write_bytes(b"later logo")
            LOGO_DISABLED_PATH.touch()
            backup_manager.restore_backup(clean)
            self.assertFalse(LOGO_PATH.exists())
            self.assertFalse(LOGO_DISABLED_PATH.exists())

    def test_backup_roundtrip_covers_all_four_signatures(self):
        # 3. ve 4. imza yedeğe hiç girmiyordu; geri yüklemeden sonra kalıcı
        # olarak kayboluyordu (PDF sessizce imzasız çıkıyordu).
        signatures = (SIG1_PATH, SIG2_PATH, SIG3_PATH, SIG4_PATH)
        original = {path: f"imza-{index}".encode()
                    for index, path in enumerate(signatures, 1)}
        try:
            for path, payload in original.items():
                path.parent.mkdir(parents=True, exist_ok=True)
                path.write_bytes(payload)

            with tempfile.TemporaryDirectory(prefix="oms_sig_") as temp_dir:
                archive_path = backup_manager.create_backup(temp_dir)
                with zipfile.ZipFile(archive_path) as archive:
                    names = set(archive.namelist())
                for index in range(1, 5):
                    self.assertIn(f"signature{index}.png", names,
                                  f"{index}. imza yedeğe alınmamış")

                for path in signatures:      # yedek sonrası bozulma / silinme
                    path.write_bytes(b"sonradan-degisti")
                SIG4_PATH.unlink()

                backup_manager.restore_backup(archive_path)

                for path, payload in original.items():
                    self.assertTrue(path.exists(), f"{path.name} geri gelmedi")
                    self.assertEqual(path.read_bytes(), payload,
                                     f"{path.name} içeriği yedekle aynı değil")
        finally:
            for path in signatures:
                if path.exists():
                    path.unlink()

    def test_localized_numbers_and_versions(self):
        self.assertEqual(_parse_number("1.234,56"), 1234.56)
        self.assertEqual(_parse_number("1,234.56"), 1234.56)
        self.assertFalse(is_newer_version("v1", "v1.0"))
        self.assertTrue(is_newer_version("v1.0.1", "v1.0"))


class DialogCallContractTests(unittest.TestCase):
    """Dialog çağrı yerleri gerçek imzayla uyuşmalı.

    PDF önizleme penceresi EmailDialog'u var olmayan `to_email` parametresiyle
    çağırıyordu; buton her tıklandığında TypeError ile çöküyordu. Çağrı yerleri
    kaynaktan (AST) okunup imzaya bind edilerek bu sınıf hata yakalanır.
    """

    UI_ROOT = Path(__file__).resolve().parent.parent / "ui"

    def _call_sites(self, name: str):
        """ui/ altındaki `name(...)` çağrılarını (dosya, satır, düğüm) döndürür."""
        for path in sorted(self.UI_ROOT.rglob("*.py")):
            tree = ast.parse(path.read_text(encoding="utf-8"), filename=str(path))
            for node in ast.walk(tree):
                if (isinstance(node, ast.Call)
                        and isinstance(node.func, ast.Name)
                        and node.func.id == name):
                    yield path, node.lineno, node

    def test_email_dialog_call_sites_match_signature(self):
        from ui.dialogs.email_dialog import EmailDialog
        signature = inspect.signature(EmailDialog.__init__)

        checked = 0
        for path, lineno, node in self._call_sites("EmailDialog"):
            if any(kw.arg is None for kw in node.keywords):
                continue                      # **kwargs ile çağrı — statik kontrol dışı
            checked += 1
            keywords = {kw.arg: None for kw in node.keywords}
            try:
                signature.bind(None, *([None] * len(node.args)), **keywords)
            except TypeError as exc:
                self.fail(
                    f"{path.name}:{lineno} EmailDialog çağrısı imzayla "
                    f"uyuşmuyor -> {exc}")
        self.assertGreaterEqual(checked, 3, "EmailDialog çağrı yerleri bulunamadı")


if __name__ == "__main__":
    unittest.main(verbosity=2)
