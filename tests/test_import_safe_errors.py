"""R10c-4 — içe/dışa aktarma güvenli hata sözleşmesi.

Kapsam: `ui/utils/excel_import.py`

Sözleşme:
  * Ham istisna metni, traceback, SQL, mutlak yol, firma adı, ürün kodu ve
    teklif numarası hiçbir kullanıcı mesajına, `errors` listesine veya loga
    girmez; `exc_info` kullanılmaz.
  * Her teknik istisna `operation_error.logla` ile **en fazla bir kez**
    güvenli loglanır; güvenli `kayit_id` olarak yalnız satır SIRASI kullanılır.
  * Beklenen fallback (kodlama/ayraç denemesi) hata değildir: kullanıcıya
    hata gösterilmez, loga ham istisna veya dosya adı yazılmaz.
  * Aşama ayrımı (invariant 18b): sonraki aşamanın hatası, tamamlanmış
    aşamayı "yapılamadı" gibi anlatamaz; ilerleme penceresi HER yolda kapanır.
  * Normal doğrulama sonuçları (eksik alan, mükerrer kayıt, sayfa adı)
    teknik hataya dönüşmez.

Gerçek kullanıcı DB'si, belgeleri ve dış dosyaları KULLANILMAZ: tüm dosyalar
`TemporaryDirectory` altında üretilir, DB `tests/conftest.py` izolasyonundadır.
"""
import inspect
import logging
import os

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

import tempfile
import unittest
from pathlib import Path
from unittest import mock

from PySide6.QtWidgets import QApplication, QMessageBox

import ui.utils.excel_import as xi
from ui.utils import operation_error as op_hata_mod

# Hiçbir yere sızmaması gereken içerik
FIRMA = "Gizli Müşteri A.Ş."
URUN_KODU = "GIZLI-KOD-001"
TEKLIF_NO = "TKF-2026-0777"
GIZLI = (f"no such table: offers SELECT * FROM customers WHERE company_name='{FIRMA}' "
         "C:\\Users\\Universe\\AppData\\Local\\OfferManagementSystem\\data\\database.db")
SIZINTI = ("SELECT * FROM customers", "no such table", "C:\\Users\\Universe",
           "database.db", FIRMA, URUN_KODU, TEKLIF_NO, "Traceback")


def _hata(sinif=RuntimeError, metin=None):
    try:
        raise sinif(metin if metin is not None else GIZLI)
    except Exception as exc:                                   # noqa: BLE001
        return exc


class _LogYakala(logging.Handler):
    def __init__(self):
        super().__init__()
        self.kayitlar = []

    def emit(self, k):
        metin = str(k.getMessage())
        if k.exc_info:
            import traceback
            metin += "".join(traceback.format_exception(*k.exc_info))
        self.kayitlar.append(metin)

    @property
    def birlesik(self):
        return "\n".join(self.kayitlar)


class _Temel(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.app = QApplication.instance() or QApplication([])

    def setUp(self):
        self.log = _LogYakala()
        kok = logging.getLogger()
        kok.addHandler(self.log)
        self._eski = kok.level
        kok.setLevel(logging.DEBUG)
        self.addCleanup(lambda: (kok.removeHandler(self.log),
                                 kok.setLevel(self._eski)))

        self.tmp = tempfile.TemporaryDirectory(prefix="oms_import_")
        self.addCleanup(self.tmp.cleanup)
        self.kok = Path(self.tmp.name)

        # Tüm kullanıcı kutuları yakalanır; hiçbiri gerçekten açılmaz.
        self.kutular = []
        for ad in ("warning", "information", "critical"):
            mock.patch.object(
                QMessageBox, ad,
                staticmethod(lambda p, b, m, *a, **k:
                             (self.kutular.append((b, m)),
                              QMessageBox.StandardButton.Ok)[1])).start()
        mock.patch.object(QMessageBox, "exec",
                          lambda kutu, *a, **k: (
                              self.kutular.append(
                                  (kutu.windowTitle(), kutu.text() or "")),
                              QMessageBox.StandardButton.Ok)[1]).start()
        self.addCleanup(mock.patch.stopall)

        # İlerleme penceresi: gerçek Qt penceresi açılmaz, kapanış SAYILIR.
        self.prog_kapanis = []
        self.prog_acilis = []
        test = self

        class _SahteProgress:
            def __init__(self, parent, label):
                test.prog_acilis.append(label)
                self.kapandi = False

            def set_label(self, t):
                pass

            def __call__(self, *a):
                pass

            def close(self):
                self.kapandi = True
                test.prog_kapanis.append(True)

        mock.patch.object(xi, "_ImportProgress", _SahteProgress).start()

        from core.app_paths import LOG_DIR
        LOG_DIR.mkdir(parents=True, exist_ok=True)

    # ── yardımcılar ─────────────────────────────────────────────────────
    def _log_sayaci(self):
        """`op_hata.logla` çağrılarını (islem, kayit_id) olarak sayar."""
        cagrilar = []
        gercek = op_hata_mod.logla

        def _sar(exc, islem, kayit_id=None):
            cagrilar.append((type(exc).__name__, islem, kayit_id))
            return gercek(exc, islem, kayit_id=kayit_id)

        mock.patch.object(op_hata_mod, "logla", _sar).start()
        return cagrilar

    def _metinler(self):
        return "\n".join(f"{b}\n{m}" for b, m in self.kutular)

    def _sizinti_yok(self, nerede="", ekstra=()):
        havuz = self._metinler() + "\n" + self.log.birlesik + "\n" + "\n".join(
            str(x) for x in getattr(self, "_ek_havuz", []))
        for parca in SIZINTI + tuple(ekstra):
            self.assertNotIn(parca, havuz, f"sızıntı{nerede}: {parca}")
        self.assertNotIn("exc_info", self.log.birlesik)

    def _kaydet(self, ad: str, icerik: bytes) -> Path:
        yol = self.kok / ad
        yol.write_bytes(icerik)
        return yol

    def _xlsx(self, ad="veri.xlsx", sayfalar=None) -> Path:
        """Gerçek bir XLSX üretir (openpyxl yoksa test atlanır)."""
        try:
            import openpyxl
        except ImportError:                                    # pragma: no cover
            self.skipTest("openpyxl yok")
        sayfalar = sayfalar or {"Müşteriler": [["Firma Adı"], ["Alfa Ltd."]]}
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for ad_s, satirlar in sayfalar.items():
            ws = wb.create_sheet(ad_s)
            for s in satirlar:
                ws.append(s)
        yol = self.kok / ad
        wb.save(str(yol))
        wb.close()
        return yol

    def _dosya_sec(self, yol):
        return mock.patch.object(
            xi.QFileDialog, "getOpenFileName",
            staticmethod(lambda *a, **k: (str(yol), "")))

    def _kaydet_sec(self, yol):
        return mock.patch.object(
            xi.QFileDialog, "getSaveFileName",
            staticmethod(lambda *a, **k: (str(yol), "")))

    def _onayla(self, kabul=True):
        """Onay kutusunda 'Aktar' / 'İptal' seçimini taklit eder."""
        def _exec(kutu, *a, **k):
            self.kutular.append((kutu.windowTitle(), kutu.text() or ""))
            dugmeler = kutu.buttons()
            hedef = None
            for b in dugmeler:
                rol = kutu.buttonRole(b)
                istenen = (QMessageBox.ButtonRole.AcceptRole if kabul
                           else QMessageBox.ButtonRole.RejectRole)
                if rol == istenen:
                    hedef = b
                    break
            if hedef is not None:
                kutu.setProperty("_secilen", True)
                hedef.click()
            return QMessageBox.StandardButton.Ok
        return mock.patch.object(QMessageBox, "exec", _exec)


# ── 1. Dosya okuma ──────────────────────────────────────────────────────

class DosyaOkumaTests(_Temel):

    def test_read_file_ham_istisna_dondurmez(self):
        cagrilar = self._log_sayaci()
        yol = self._xlsx()
        with mock.patch.object(xi, "_sayfa_adaylari", side_effect=_hata()):
            rows, err = xi._read_file(str(yol), import_type="customers")
        self.assertEqual(rows, [])
        self.assertTrue(err, "hata mesajı üretilmedi")
        self._ek_havuz = [err]
        self._sizinti_yok(" (_read_file)")
        self.assertEqual(len(cagrilar), 1, f"tek güvenli log değil: {cagrilar}")

    def test_read_xlsx_sheets_ham_istisna_dondurmez(self):
        cagrilar = self._log_sayaci()
        bozuk = self._kaydet("bozuk.xlsx", b"bu bir xlsx degil")
        sheets, err = xi._read_xlsx_sheets(str(bozuk))
        self.assertEqual(sheets, {})
        self.assertTrue(err)
        self._ek_havuz = [err]
        self._sizinti_yok(" (_read_xlsx_sheets)")
        self.assertEqual(len(cagrilar), 1, f"tek güvenli log değil: {cagrilar}")

    def test_okuma_hatasi_mesaji_sabit(self):
        """İki farklı teknik neden AYNI sabit kullanıcı metnini üretir."""
        yol = self._xlsx()
        with mock.patch.object(xi, "_sayfa_adaylari", side_effect=_hata(OSError)):
            _r1, e1 = xi._read_file(str(yol), import_type="customers")
        with mock.patch.object(xi, "_sayfa_adaylari", side_effect=_hata(ValueError)):
            _r2, e2 = xi._read_file(str(yol), import_type="customers")
        self.assertEqual(e1, e2, "mesaj teknik nedene göre değişiyor")

    def test_sayfa_secimi_iptal_isareti_korunur(self):
        """Gerçek kullanıcı iptali hâlâ SAYFA_SECIMI_IPTAL üretir."""
        yol = self._xlsx(sayfalar={
            "Sayfa1": [["Firma Adı"], ["A"]],
            "Sayfa2": [["Firma Adı"], ["B"]],
        })
        with mock.patch.object(xi, "_sayfa_sordur",
                               staticmethod(lambda *a, **k: None)):
            rows, err = xi._read_file(str(yol), import_type="customers")
        self.assertEqual(rows, [])
        self.assertEqual(err, xi.SAYFA_SECIMI_IPTAL)

    def test_onceden_secim_hatasi_ikinci_kez_loglanmaz(self):
        """`_sayfa_sec_onceden` sessiz döner; log `_read_file`'da TEK kez atılır."""
        cagrilar = self._log_sayaci()
        bozuk = self._kaydet("bozuk.xlsx", b"xlsx degil")
        secilen, hata = xi._sayfa_sec_onceden(str(bozuk), "customers", None)
        self.assertIsNone(secilen)
        self.assertEqual(hata, "")
        self.assertEqual(len(cagrilar), 0,
                         f"ön seçim aşaması ayrıca logladı: {cagrilar}")
        xi._read_file(str(bozuk), import_type="customers")
        self.assertEqual(len(cagrilar), 1, f"tek log değil: {cagrilar}")

    def test_progress_okuma_hatasinda_kapanir(self):
        yol = self._xlsx()
        with self._dosya_sec(yol), \
             mock.patch.object(xi, "_read_file",
                               staticmethod(lambda *a, **k: ([], "Dosya okunamadı."))):
            self.assertIs(xi.run_import_flow(None, "customers"), False)
        self.assertEqual(len(self.prog_kapanis), 1, "ilerleme penceresi kapanmadı")

    def test_sayfa_secimi_ilerleme_penceresinden_once(self):
        """İptal edilen sayfa seçiminde ilerleme penceresi HİÇ açılmaz."""
        yol = self._xlsx()
        with self._dosya_sec(yol), \
             mock.patch.object(xi, "_sayfa_sec_onceden", staticmethod(
                 lambda *a, **k: (None, xi.SAYFA_SECIMI_IPTAL))):
            self.assertIs(xi.run_import_flow(None, "customers"), False)
        self.assertEqual(self.prog_acilis, [], "iptalde ilerleme penceresi açıldı")
        self.assertEqual(self.kutular, [], "iptalde kutu gösterildi")


# ── 2. CSV güvenli logları ──────────────────────────────────────────────

class CsvGuvenliLogTests(_Temel):

    def test_csv_acilamadi_tek_guvenli_log(self):
        cagrilar = self._log_sayaci()
        yol = self._kaydet("Gizli Müşteri A.Ş..csv", b"Firma\nA\n")
        with mock.patch.object(Path, "read_bytes", side_effect=_hata(OSError)):
            rows, err = xi._read_csv(yol)
        self.assertEqual(rows, [])
        self.assertTrue(err)
        self._ek_havuz = [err]
        self._sizinti_yok(" (csv aç)", ekstra=("Gizli Müşteri",))
        self.assertEqual(len(cagrilar), 1, f"tek güvenli log değil: {cagrilar}")

    def test_ayrac_fallback_kullaniciya_hata_gostermez(self):
        yol = self._kaydet("veri.csv", "Firma Adı\nAlfa Ltd.\nBeta A.Ş.\n"
                           .encode("utf-8"))
        rows, err = xi._read_csv(yol)
        self.assertEqual(err, "", "okunabilir tek sütunlu dosya reddedildi")
        self.assertEqual(len(rows), 2)

    def test_ayrac_fallback_logu_ham_veri_tasimaz(self):
        yol = self._kaydet(f"{FIRMA}.csv",
                           "Firma Adı\nAlfa Ltd.\n".encode("utf-8"))
        xi._read_csv(yol)
        self._sizinti_yok(" (ayraç fallback)")
        for parca in ("csv.Error", "Could not determine"):
            self.assertNotIn(parca, self.log.birlesik,
                             f"ayraç fallback logunda ham istisna: {parca}")

    def test_tum_denemeler_basarisizsa_tek_guvenli_log(self):
        cagrilar = self._log_sayaci()
        yol = self._kaydet(f"{FIRMA}.csv", "Firma\nA\n".encode("utf-8"))
        with mock.patch.object(xi.csv, "DictReader",
                               side_effect=xi.csv.Error(GIZLI)):
            rows, err = xi._read_csv(yol)
        self.assertEqual(rows, [])
        self.assertTrue(err, "son anlamlı hata kayboldu")
        self._ek_havuz = [err]
        self._sizinti_yok(" (csv ayrıştır)")
        self.assertEqual(len(cagrilar), 1,
                         f"son anlamlı neden tam bir kez loglanmadı: {cagrilar}")

    def test_gercek_bos_dosya_hata_degil(self):
        yol = self._kaydet("bos.csv", b"   \n\n")
        rows, err = xi._read_csv(yol)
        self.assertEqual(rows, [])
        self.assertEqual(err, "", "boş dosya hata sayıldı")

    def test_ikili_icerik_sabit_hata(self):
        yol = self._kaydet("ikili.csv", bytes(range(256)) * 20)
        rows, err = xi._read_csv(yol)
        self.assertEqual(rows, [])
        self.assertEqual(err, xi._CSV_OKUMA_HATASI)

    def test_csv_loglarinda_dosya_adi_yok(self):
        """Dosya adı kullanıcı verisi taşıyabilir; loga girmemeli."""
        for satir in inspect.getsource(xi._read_csv).splitlines():
            if "logger." in satir:
                self.assertNotIn("path.name", satir, f"dosya adı loglanıyor: {satir}")


# ── 3. Müşteri ve ürün satır hataları ───────────────────────────────────

class SatirHatasiTests(_Temel):

    def _musteri_satirlari(self, n=3):
        return [{"company_name": f"{FIRMA} {i}"} for i in range(n)]

    def _urun_satirlari(self, n=3):
        return [{"product_code": f"{URUN_KODU}-{i}", "product_name": f"Ürün {i}",
                 "price": "10", "stock": "1"} for i in range(n)]

    def _patlayan_conn(self, patlayan_indeksler):
        """Belirtilen sıradaki `execute` çağrılarında hata veren sahte conn."""
        durum = {"n": 0}
        gercek_tx = xi_db_transaction()

        class _Conn:
            def __init__(self, ic):
                self._ic = ic

            def execute(self, *a, **k):
                i = durum["n"]
                durum["n"] += 1
                if i in patlayan_indeksler:
                    raise RuntimeError(GIZLI)
                return self._ic.execute(*a, **k)

            def __getattr__(self, ad):
                return getattr(self._ic, ad)

        return gercek_tx, _Conn

    def test_musteri_satir_hatasi_guvenli(self):
        cagrilar = self._log_sayaci()
        rows = self._musteri_satirlari(3)
        with _sarmalayan_transaction(patlayanlar={0}):
            added, updated, skipped, errors = xi._perform_import(
                "customers", rows, False)
        self._ek_havuz = list(errors)
        self._sizinti_yok(" (müşteri satırı)")
        self.assertEqual(len(errors), 1)
        self.assertRegex(errors[0], r"Satır\s*1")
        self.assertEqual(added, 2, "bir satırın hatası sonrakileri engelledi")
        self.assertEqual(len(cagrilar), 1, f"tek güvenli log değil: {cagrilar}")
        self.assertEqual(cagrilar[0][2], 1, "güvenli kayit_id satır sırası değil")

    def test_urun_satir_hatasi_guvenli(self):
        cagrilar = self._log_sayaci()
        rows = self._urun_satirlari(3)
        with _sarmalayan_transaction(patlayanlar={1}):
            added, updated, skipped, errors = xi._perform_import(
                "products", rows, False)
        self._ek_havuz = list(errors)
        self._sizinti_yok(" (ürün satırı)")
        self.assertEqual(len(errors), 1)
        self.assertRegex(errors[0], r"Satır\s*2")
        self.assertEqual(added, 2, "bir satırın hatası sonrakileri engelledi")
        self.assertEqual(len(cagrilar), 1, f"tek güvenli log değil: {cagrilar}")
        self.assertEqual(cagrilar[0][2], 2)

    def test_satir_disi_fatal_hata_cagirana_tasinir(self):
        """Commit gibi satır dışı hata yutulmaz; çağıran güvenle ele alır."""
        rows = self._musteri_satirlari(1)
        with mock.patch("database.db_manager.get_db",
                        side_effect=_hata(RuntimeError)):
            with self.assertRaises(Exception):
                xi._perform_import("customers", rows, False)

    def test_yazma_hatasinda_progress_kapanir_ve_false_doner(self):
        yol = self._xlsx()
        with self._dosya_sec(yol), \
             mock.patch.object(xi, "_read_file", staticmethod(
                 lambda *a, **k: ([{"Firma Adı": "Alfa"}], ""))), \
             mock.patch.object(xi, "_validate_rows", staticmethod(
                 lambda *a, **k: ([{"company_name": "Alfa"}], [], []))), \
             mock.patch.object(xi, "_perform_import", side_effect=_hata()), \
             self._onayla(True):
            sonuc = xi.run_import_flow(None, "customers")
        self.assertIs(sonuc, False, "teknik hata sonrası yanlış başarı döndü")
        self.assertEqual(len(self.prog_kapanis), 2,
                         "ilerleme penceresi yazma hatasında kapanmadı")
        self._sizinti_yok(" (yazma hatası)")
        self.assertNotRegex(self._metinler(), r"(?i)aktarıldı|Tamamlandı")


# ── 4. Kategori oluşturma ───────────────────────────────────────────────

class KategoriTests(_Temel):

    def test_kategori_hatasi_guvenli_ve_tekrarsiz(self):
        cagrilar = self._log_sayaci()
        rows = [{"product_code": f"KAT-A-{i}", "product_name": f"Ü{i}",
                 "price": "1", "stock": "1", "category": "Gizli Kategori"}
                for i in range(3)]
        with mock.patch("services.category_service.CategoryService.add",
                        side_effect=_hata()):
            added, updated, skipped, errors = xi._perform_import(
                "products", rows, False)
        self.assertEqual(added, 3, "kategori hatası tamamlanmış ürün kaydını iptal etti")
        self._ek_havuz = list(errors)
        self._sizinti_yok(" (kategori)", ekstra=("Gizli Kategori",))
        self.assertEqual(len(cagrilar), 1,
                         f"kategori hatası tekrar tekrar loglandı: {cagrilar}")
        self.assertLessEqual(len(errors), 1,
                             f"yinelenen kategori hata satırları: {errors}")

    def test_kategori_hatasi_kullaniciya_bildirilir(self):
        rows = [{"product_code": "KAT-B-1", "product_name": "Ü1",
                 "price": "1", "stock": "1", "category": "Gizli Kategori"}]
        with mock.patch("services.category_service.CategoryService.add",
                        side_effect=_hata()):
            added, _u, _s, errors = xi._perform_import("products", rows, False)
        self.assertEqual(added, 1)
        self.assertTrue(errors, "kategori aşamasındaki sorun kullanıcıya söylenmedi")
        self.assertRegex(" ".join(errors), r"(?i)kategori")


# ── 5. Teklif kayıt hataları ────────────────────────────────────────────

class TeklifHatasiTests(_Temel):

    def _grup(self, no):
        return {"offer_no": no, "company_name": FIRMA, "contact_person": "",
                "address": "", "phone": "", "email": "", "date": "",
                "currency": "EUR", "status": "Beklemede", "validity": "",
                "validity_note": "", "payment_term": "", "discount_percent": 0,
                "items": [{"product_code": URUN_KODU, "product_name": "Ü",
                           "description": "", "quantity": 1.0, "unit": "Adet",
                           "delivery_time": "", "unit_price": 10.0}]}

    def test_teklif_hatasi_guvenli_ve_sonrakini_engellemez(self):
        cagrilar = self._log_sayaci()
        gruplar = [self._grup(TEKLIF_NO), self._grup("TKF-2"), self._grup("TKF-3")]
        cagri = {"n": 0}
        gercek = None

        def _save(self_svc, offer, **k):
            cagri["n"] += 1
            if cagri["n"] == 1:
                raise RuntimeError(GIZLI)
            return 1

        with mock.patch("services.offer_service.OfferService.save", _save):
            added, errors = xi._perform_offer_import(gruplar)
        self._ek_havuz = list(errors)
        self._sizinti_yok(" (teklif)")
        self.assertEqual(added, 2, "bir grubun hatası sonrakileri engelledi")
        self.assertEqual(len(errors), 1)
        self.assertEqual(len(cagrilar), 1, f"tek güvenli log değil: {cagrilar}")
        self.assertIn(cagrilar[0][2], (1, "1"),
                      "güvenli kayit_id grup sırası değil")

    def test_teklif_dogrulama_mesajlari_korunur(self):
        """Eksik zorunlu alan ve mükerrer kayıt TEKNİK hata değildir."""
        ham = [{"Teklif No": "", "Firma": "A"},
               {"Teklif No": "T1", "Firma": ""}]
        yeni, dups, invalid = xi._validate_offer_rows(ham)
        birlesik = " ".join(invalid)
        self.assertRegex(birlesik, r"(?i)teklif no eksik")
        self.assertRegex(birlesik, r"(?i)firma ad. eksik")


# ── 6. Tekli import üst seviye aşamaları ────────────────────────────────

class AkisAsamaTests(_Temel):

    def test_kullanici_iptalinde_yan_etki_yok(self):
        cagrilar = self._log_sayaci()
        yol = self._xlsx()
        yazildi = []
        with self._dosya_sec(yol), \
             mock.patch.object(xi, "_read_file", staticmethod(
                 lambda *a, **k: ([{"Firma Adı": "Alfa"}], ""))), \
             mock.patch.object(xi, "_validate_rows", staticmethod(
                 lambda *a, **k: ([{"company_name": "Alfa"}], [], []))), \
             mock.patch.object(xi, "_perform_import",
                               side_effect=lambda *a, **k: yazildi.append(1)), \
             self._onayla(False):
            self.assertIs(xi.run_import_flow(None, "customers"), False)
        self.assertEqual(yazildi, [], "iptalde DB'ye yazıldı")
        self.assertEqual(len(cagrilar), 0, "iptalde log atıldı")
        for _b, m in self.kutular:
            self.assertNotRegex(m or "", r"(?i)hata")

    def test_dogrulama_teknik_hatasi_guvenli(self):
        cagrilar = self._log_sayaci()
        yol = self._xlsx()
        with self._dosya_sec(yol), \
             mock.patch.object(xi, "_read_file", staticmethod(
                 lambda *a, **k: ([{"Firma Adı": "Alfa"}], ""))), \
             mock.patch.object(xi, "_validate_rows", side_effect=_hata()):
            self.assertIs(xi.run_import_flow(None, "customers"), False)
        self.assertEqual(len(self.prog_kapanis), 1,
                         "doğrulama hatasında ilerleme penceresi kapanmadı")
        self.assertEqual(len(cagrilar), 1, f"tek güvenli log değil: {cagrilar}")
        self._sizinti_yok(" (doğrulama)")


# ── 7. Tümünü İçe Aktar aşama doğruluğu ─────────────────────────────────

class TumunuAktarTests(_Temel):

    def _sayfalar(self):
        return {
            "Müşteriler": [{"Firma Adı": "Alfa Ltd."}],
            "Ürünler": [{"Ürün Kodu": "P1", "Ürün Adı": "Ü1", "Fiyat": "10"}],
        }

    def _akis(self, perform_side_effect):
        yol = self._xlsx()
        return (self._dosya_sec(yol),
                mock.patch.object(xi, "_read_xlsx_sheets", staticmethod(
                    lambda *a, **k: (self._sayfalar(), ""))),
                mock.patch.object(xi, "_validate_rows", staticmethod(
                    lambda tur, rows, **k: ([{"company_name": "A"}]
                                            if tur == "customers" else
                                            [{"product_code": "P1",
                                              "product_name": "Ü1"}], [], []))),
                mock.patch.object(xi, "_validate_offer_rows", staticmethod(
                    lambda *a, **k: ([], [], []))),
                mock.patch.object(xi, "_perform_import", perform_side_effect),
                self._onayla(True))

    def test_sonraki_asama_hatasi_oncekini_inkar_etmez(self):
        cagrilar = self._log_sayaci()

        def _perform(tur, rows, upd, progress=None, stage_state=None):
            if tur == "products":
                raise _hata()
            return 5, 0, 0, []

        a, b, c, d, e, f = self._akis(_perform)
        with a, b, c, d, e, f:
            sonuc = xi.run_import_all_flow(None)
        metin = self._metinler()
        self.assertRegex(metin, r"Müşteri:\s*5 eklendi",
                         "tamamlanmış müşteri aşaması inkâr edildi")
        self.assertRegex(metin, r"(?i)ürün.*(aktarılamadı|kaydedilemedi|başarısız)",
                         "başarısız aşama ayrı gösterilmedi")
        self.assertIs(sonuc, True, "DB değişti ama cache yenilemesi tetiklenmedi")
        self.assertEqual(len(cagrilar), 1, f"tek güvenli log değil: {cagrilar}")
        self._sizinti_yok(" (tümünü aktar)")
        self.assertEqual(len(self.prog_kapanis), 2, "ilerleme penceresi kapanmadı")

    def test_hicbir_asama_tamamlanmazsa_false(self):
        def _perform(tur, rows, upd, progress=None, stage_state=None):
            raise _hata()

        a, b, c, d, e, f = self._akis(_perform)
        with a, b, c, d, e, f:
            sonuc = xi.run_import_all_flow(None)
        self.assertIs(sonuc, False,
                      "hiçbir aşama tamamlanmadığı hâlde True döndü")
        self.assertEqual(len(self.prog_kapanis), 2, "ilerleme penceresi kapanmadı")

    def test_asama_hatasinda_sonraki_asama_denenir(self):
        """Aşamalar bağımsız transaction'lardır; biri düşünce diğeri atlanmaz."""
        gorulen = []

        def _perform(tur, rows, upd, progress=None, stage_state=None):
            gorulen.append(tur)
            if tur == "customers":
                raise _hata()
            return 3, 0, 0, []

        a, b, c, d, e, f = self._akis(_perform)
        with a, b, c, d, e, f:
            sonuc = xi.run_import_all_flow(None)
        self.assertEqual(gorulen, ["customers", "products"],
                         "önceki aşamanın hatası sonrakini engelledi")
        self.assertIs(sonuc, True)
        self.assertRegex(self._metinler(), r"Ürün:\s*3 eklendi")


# ── 8. Dışa aktarma catch'leri ──────────────────────────────────────────

class ExportTests(_Temel):

    def test_veri_okuma_hatasi_guvenli(self):
        cagrilar = self._log_sayaci()
        with mock.patch("services.customer_service.CustomerService.get_all",
                        side_effect=_hata()):
            xi.export_data_interactive(None, "customers")
        self._sizinti_yok(" (export oku)")
        self.assertEqual(len(cagrilar), 1, f"tek güvenli log değil: {cagrilar}")

    def test_dosya_yazma_hatasi_okuma_basarisini_yanlis_anlatmaz(self):
        cagrilar = self._log_sayaci()
        hedef = self.kok / "cikti.xlsx"
        with self._kaydet_sec(hedef), \
             mock.patch("services.customer_service.CustomerService.get_all",
                        staticmethod(lambda: [])), \
             mock.patch("services.export_service.export_customers_excel",
                        side_effect=_hata()):
            xi.export_data_interactive(None, "customers")
        metin = self._metinler()
        self.assertNotRegex(metin, r"(?i)veriler okunamadı",
                            "yazma hatası okuma hatası gibi anlatıldı")
        self._sizinti_yok(" (export yaz)")
        self.assertEqual(len(cagrilar), 1, f"tek güvenli log değil: {cagrilar}")

    def test_iptalde_log_ve_hata_yok(self):
        cagrilar = self._log_sayaci()
        with mock.patch.object(xi.QFileDialog, "getSaveFileName",
                               staticmethod(lambda *a, **k: ("", ""))), \
             mock.patch("services.customer_service.CustomerService.get_all",
                        staticmethod(lambda: [])):
            xi.export_data_interactive(None, "customers")
        self.assertEqual(cagrilar, [], "iptalde log atıldı")
        self.assertEqual(self.kutular, [], "iptalde kutu gösterildi")

    def test_tumunu_disa_aktar_okuma_hatasi_guvenli(self):
        cagrilar = self._log_sayaci()
        with mock.patch("services.customer_service.CustomerService.get_all",
                        side_effect=_hata()):
            xi.export_all_interactive(None)
        self._sizinti_yok(" (export all oku)")
        self.assertEqual(len(cagrilar), 1, f"tek güvenli log değil: {cagrilar}")

    def test_tumunu_disa_aktar_yazma_hatasi_guvenli(self):
        cagrilar = self._log_sayaci()
        hedef = self.kok / "hepsi.xlsx"
        with self._kaydet_sec(hedef), \
             mock.patch("services.export_service.export_all_excel",
                        side_effect=_hata()):
            xi.export_all_interactive(None)
        self._sizinti_yok(" (export all yaz)")
        self.assertEqual(len(cagrilar), 1, f"tek güvenli log değil: {cagrilar}")


# ── 10. Farklı kategori hataları ────────────────────────────────────────

class KategoriCoklulukTests(_Temel):

    def test_iki_farkli_kategori_ayri_ayri_loglanir(self):
        """Aynı kategori TEK kez denenir; FARKLI kategoriler ayrı istisnadır."""
        cagrilar = self._log_sayaci()
        rows = [
            {"product_code": "MK-1", "product_name": "Ü1", "price": "1",
             "stock": "1", "category": "Gizli Kategori A"},
            {"product_code": "MK-2", "product_name": "Ü2", "price": "1",
             "stock": "1", "category": "Gizli Kategori A"},
            {"product_code": "MK-3", "product_name": "Ü3", "price": "1",
             "stock": "1", "category": "Gizli Kategori B"},
        ]
        denenen = []

        def _add(self_svc, kategori):
            denenen.append(kategori.name)
            raise RuntimeError(GIZLI)

        with mock.patch("services.category_service.CategoryService.add", _add):
            added, _u, _s, errors = xi._perform_import("products", rows, False)
        self.assertEqual(denenen, ["Gizli Kategori A", "Gizli Kategori B"],
                         "aynı kategori yeniden denendi veya biri atlandı")
        self.assertEqual(len(cagrilar), 2,
                         f"her farklı kategori hatası bir kez loglanmadı: {cagrilar}")
        self.assertEqual(added, 3, "kategori hatası ürün kaydını iptal etti")
        self.assertEqual([e for e in errors if "kategori" in e.lower()],
                         [xi.KATEGORI_UYARISI],
                         f"tek toplu kategori uyarısı değil: {errors}")
        self._ek_havuz = list(errors)
        self._sizinti_yok(" (çoklu kategori)",
                          ekstra=("Gizli Kategori A", "Gizli Kategori B"))


# ── 11. Tümünü aktar doğrulama koruması ─────────────────────────────────

class TumunuDogrulamaTests(_Temel):

    def _sayfalar(self):
        return {
            "Müşteriler": [{"Firma Adı": "Alfa Ltd."}],
            "Ürünler": [{"Ürün Kodu": "P1", "Ürün Adı": "Ü1"}],
            "Teklifler": [{"Teklif No": "T1", "Firma": "Alfa"}],
        }

    def _calistir(self, **patchler):
        yol = self._xlsx()
        yazimlar = []
        with self._dosya_sec(yol), \
             mock.patch.object(xi, "_read_xlsx_sheets", staticmethod(
                 lambda *a, **k: (self._sayfalar(), ""))), \
             mock.patch.object(xi, "_perform_import",
                               side_effect=lambda *a, **k: yazimlar.append("p")), \
             mock.patch.object(xi, "_perform_offer_import",
                               side_effect=lambda *a, **k: yazimlar.append("o")), \
             mock.patch.object(xi, "_validate_rows", staticmethod(
                 patchler.get("rows", lambda t, r, **k: ([], [], [])))), \
             mock.patch.object(xi, "_validate_offer_rows", staticmethod(
                 patchler.get("offers", lambda *a, **k: ([], [], [])))):
            sonuc = xi.run_import_all_flow(None)
        return sonuc, yazimlar

    def _dogrula(self, sonuc, yazimlar, cagrilar, nerede):
        self.assertIs(sonuc, False, f"{nerede}: teknik hatada True döndü")
        self.assertEqual(yazimlar, [], f"{nerede}: doğrulama hatasında DB'ye yazıldı")
        self.assertEqual(len(cagrilar), 1,
                         f"{nerede}: tek güvenli log değil: {cagrilar}")
        self.assertEqual(len(self.prog_kapanis), 1,
                         f"{nerede}: ilerleme penceresi tam bir kez kapanmadı")
        self.assertTrue(self.kutular, f"{nerede}: kullanıcıya hiçbir şey söylenmedi")
        self._sizinti_yok(f" ({nerede})")

    def test_musteri_dogrulama_teknik_hatasi(self):
        cagrilar = self._log_sayaci()

        def _rows(tur, r, **k):
            if tur == "customers":
                raise _hata()
            return [], [], []

        sonuc, yazimlar = self._calistir(rows=_rows)
        self._dogrula(sonuc, yazimlar, cagrilar, "müşteri doğrulama")

    def test_urun_dogrulama_teknik_hatasi(self):
        cagrilar = self._log_sayaci()

        def _rows(tur, r, **k):
            if tur == "products":
                raise _hata()
            return [], [], []

        sonuc, yazimlar = self._calistir(rows=_rows)
        self._dogrula(sonuc, yazimlar, cagrilar, "ürün doğrulama")

    def test_teklif_dogrulama_teknik_hatasi(self):
        cagrilar = self._log_sayaci()

        def _offers(*a, **k):
            raise _hata()

        sonuc, yazimlar = self._calistir(offers=_offers)
        self._dogrula(sonuc, yazimlar, cagrilar, "teklif doğrulama")

    def test_normal_dogrulama_mesajlari_degismedi(self):
        """Eksik alan / mükerrer / sayfa adı yolları teknik hataya dönüşmedi."""
        yol = self._xlsx()
        with self._dosya_sec(yol), \
             mock.patch.object(xi, "_read_xlsx_sheets", staticmethod(
                 lambda *a, **k: ({"Baska": []}, ""))):
            self.assertIs(xi.run_import_all_flow(None), False)
        self.assertRegex(self._metinler(), r"(?i)sayfa bulunamadı")


# ── 12. Teklif tekli akış üst sınırı ────────────────────────────────────

class TeklifAkisSiniriTests(_Temel):

    def test_offer_akis_teknik_hatasi_guvenli(self):
        cagrilar = self._log_sayaci()
        yol = self._xlsx()
        with self._dosya_sec(yol), \
             mock.patch.object(xi, "_read_file", staticmethod(
                 lambda *a, **k: ([{"Teklif No": "T1"}], ""))), \
             mock.patch.object(xi, "_run_offer_import_flow", side_effect=_hata()):
            sonuc = xi.run_import_flow(None, "offers")
        self.assertIs(sonuc, False, "teknik hata sonrası yanlış başarı döndü")
        self.assertEqual(len(cagrilar), 1, f"tek güvenli log değil: {cagrilar}")
        self.assertEqual(len(self.prog_kapanis), 1,
                         "okuma ilerleme penceresi tam bir kez kapanmadı")
        self.assertNotRegex(self._metinler(), r"(?i)Tamamlandı")
        self._sizinti_yok(" (teklif akışı)")


# ── 13. Kategori yazımı + ürün transaction fatal hatası ─────────────────

class AsamaDurumuTests(_Temel):

    def _urun_akisi(self, perform):
        yol = self._xlsx()
        return (self._dosya_sec(yol),
                mock.patch.object(xi, "_read_file", staticmethod(
                    lambda *a, **k: ([{"Ürün Kodu": "P1"}], ""))),
                mock.patch.object(xi, "_validate_rows", staticmethod(
                    lambda *a, **k: ([{"product_code": "P1",
                                       "product_name": "Ü1"}], [], []))),
                mock.patch.object(xi, "_perform_import", perform),
                self._onayla(True))

    def test_kategori_yazildi_urun_dustu_true_doner(self):
        cagrilar = self._log_sayaci()

        def _perform(tur, rows, upd, progress=None, stage_state=None):
            if stage_state is not None:
                stage_state["kategori_yazildi"] = 2
            raise _hata()

        a, b, c, d, e = self._urun_akisi(_perform)
        with a, b, c, d, e:
            sonuc = xi.run_import_flow(None, "products")
        metin = self._metinler()
        self.assertIs(sonuc, True,
                      "kategori yazıldığı hâlde cache yenilemesi tetiklenmedi")
        self.assertRegex(metin, r"(?i)kategori",
                         "tamamlanan kategori aşaması inkâr edildi")
        self.assertRegex(metin, r"(?i)ürün.*(tamamlanamadı|aktarılamadı)",
                         "ürün aşamasının düştüğü söylenmedi")
        self.assertEqual(len(cagrilar), 1, f"tek güvenli log değil: {cagrilar}")
        self.assertEqual(len(self.prog_kapanis), 2, "progress kapanmadı")
        self._sizinti_yok(" (kategori + fatal)")

    def test_hicbir_yazim_yoksa_false(self):
        def _perform(tur, rows, upd, progress=None, stage_state=None):
            raise _hata()

        a, b, c, d, e = self._urun_akisi(_perform)
        with a, b, c, d, e:
            self.assertIs(xi.run_import_flow(None, "products"), False)

    def test_stage_state_kullanici_verisi_tasimaz(self):
        """Aşama durumu yalnız sayı/boolean taşır; kategori adı taşımaz."""
        gorulen = {}

        def _perform(tur, rows, upd, progress=None, stage_state=None):
            gorulen["s"] = stage_state
            return 1, 0, 0, []

        a, b, c, d, e = self._urun_akisi(_perform)
        with a, b, c, d, e:
            xi.run_import_flow(None, "products")
        durum = gorulen.get("s")
        self.assertIsInstance(durum, dict, "aşama durumu geçirilmedi")
        for k, v in durum.items():
            self.assertIsInstance(v, (int, bool),
                                  f"aşama durumu metin taşıyor: {k}={v!r}")

    def test_gercek_kategori_yazimi_stage_state_doldurur(self):
        """Gerçek `_perform_import`: yeni kategori yazımı sayı olarak bildirilir."""
        durum = {}
        rows = [{"product_code": "SS-1", "product_name": "Ü1", "price": "1",
                 "stock": "1", "category": "Yeni Kategori SS"}]
        added, _u, _s, _e = xi._perform_import("products", rows, False,
                                               stage_state=durum)
        self.assertEqual(added, 1)
        self.assertGreaterEqual(durum.get("kategori_yazildi", 0), 1,
                                "gerçek kategori yazımı bildirilmedi")

    def test_tumunu_aktarda_kategori_yazimi_yenileme_saglar(self):
        yol = self._xlsx()

        def _perform(tur, rows, upd, progress=None, stage_state=None):
            if tur == "products":
                if stage_state is not None:
                    stage_state["kategori_yazildi"] = 1
                raise _hata()
            return 0, 0, 0, []

        with self._dosya_sec(yol), \
             mock.patch.object(xi, "_read_xlsx_sheets", staticmethod(
                 lambda *a, **k: ({"Ürünler": [{"Ürün Kodu": "P1"}]}, ""))), \
             mock.patch.object(xi, "_validate_rows", staticmethod(
                 lambda *a, **k: ([{"product_code": "P1",
                                    "product_name": "Ü"}], [], []))), \
             mock.patch.object(xi, "_validate_offer_rows", staticmethod(
                 lambda *a, **k: ([], [], []))), \
             mock.patch.object(xi, "_perform_import", _perform), \
             self._onayla(True):
            sonuc = xi.run_import_all_flow(None)
        self.assertIs(sonuc, True,
                      "kategori DB'yi değiştirdiği hâlde yenileme tetiklenmedi")
        self.assertRegex(self._metinler(), r"(?i)ürün.*aktarılamadı",
                         "başarısız ürün aşaması özeti kayboldu")


# ── 14. Workbook kapatma ayrı aşamadır ──────────────────────────────────

class WorkbookKapatmaTests(_Temel):

    def _patlayan_close(self):
        """Gerçek workbook'u kullanır, yalnız `close()` hata verir."""
        import openpyxl
        gercek = openpyxl.load_workbook

        def _yukle(*a, **k):
            wb = gercek(*a, **k)
            gercek_close = wb.close

            def _close():
                # Dosya tanıtıcısı GERÇEKTEN kapatılır (Windows'ta geçici
                # klasör silinebilsin), ardından hata üretilir.
                gercek_close()
                raise OSError(GIZLI)

            wb.close = _close
            return wb

        return mock.patch.object(openpyxl, "load_workbook", _yukle)

    def test_basarili_read_file_close_hatasini_yutmaz(self):
        cagrilar = self._log_sayaci()
        yol = self._xlsx()
        with self._patlayan_close():
            rows, err = xi._read_file(str(yol), import_type="customers")
        self.assertEqual(err, "", "close hatası okuma başarısını maskeledi")
        self.assertEqual(len(rows), 1, "satırlar kayboldu")
        self.assertEqual(len(cagrilar), 1, f"tek güvenli log değil: {cagrilar}")
        self._sizinti_yok(" (read_file close)")

    def test_basarili_read_xlsx_sheets_close_hatasini_yutmaz(self):
        cagrilar = self._log_sayaci()
        yol = self._xlsx()
        with self._patlayan_close():
            sheets, err = xi._read_xlsx_sheets(str(yol))
        self.assertEqual(err, "", "close hatası okuma başarısını maskeledi")
        self.assertIn("Müşteriler", sheets)
        self.assertEqual(len(cagrilar), 1, f"tek güvenli log değil: {cagrilar}")
        self._sizinti_yok(" (read_xlsx close)")

    def test_sayfa_sec_onceden_close_hatasi_secimi_bozmaz(self):
        cagrilar = self._log_sayaci()
        yol = self._xlsx(sayfalar={
            "Sayfa1": [["Firma Adı"], ["A"]],
            "Sayfa2": [["Firma Adı"], ["B"]],
        })
        with self._patlayan_close(), \
             mock.patch.object(xi, "_sayfa_sordur",
                               staticmethod(lambda *a, **k: "Sayfa2")):
            secilen, hata = xi._sayfa_sec_onceden(str(yol), "customers", None)
        self.assertEqual(secilen, "Sayfa2", "close hatası seçimi bozdu")
        self.assertEqual(hata, "")
        self.assertEqual(len(cagrilar), 1, f"tek güvenli log değil: {cagrilar}")
        self._sizinti_yok(" (sayfa seç close)")

    def test_sayfa_sec_onceden_close_hatasi_iptali_bozmaz(self):
        yol = self._xlsx(sayfalar={
            "Sayfa1": [["Firma Adı"], ["A"]],
            "Sayfa2": [["Firma Adı"], ["B"]],
        })
        with self._patlayan_close(), \
             mock.patch.object(xi, "_sayfa_sordur",
                               staticmethod(lambda *a, **k: None)):
            secilen, hata = xi._sayfa_sec_onceden(str(yol), "customers", None)
        self.assertIsNone(secilen)
        self.assertEqual(hata, xi.SAYFA_SECIMI_IPTAL,
                         "close hatası kullanıcı iptalini bozdu")

    def test_okuma_hatasi_close_hatasini_maskelemez(self):
        """İki AYRI istisna → her biri tam bir kez; asıl okuma hatası kalır."""
        cagrilar = self._log_sayaci()
        yol = self._xlsx()
        with self._patlayan_close(), \
             mock.patch.object(xi, "_sayfa_adaylari", side_effect=_hata()):
            rows, err = xi._read_file(str(yol), import_type="customers")
        self.assertEqual(rows, [])
        self.assertEqual(err, xi.DOSYA_OKUMA_HATASI,
                         "asıl okuma hatası close hatasıyla değiştirildi")
        self.assertEqual(len(cagrilar), 2,
                         f"iki ayrı istisna birer kez loglanmadı: {cagrilar}")
        self._ek_havuz = [err]
        self._sizinti_yok(" (okuma + close)")


# ── 15. Hata yolunda workbook kapanması ─────────────────────────────────

class WorkbookHataYoluTests(_Temel):

    def _sayac_close(self):
        """Gerçek workbook; `close()` sayılır ve hata verir."""
        import openpyxl
        gercek = openpyxl.load_workbook
        sayac = []

        def _yukle(*a, **k):
            wb = gercek(*a, **k)
            gercek_close = wb.close

            def _close():
                sayac.append(1)
                gercek_close()                 # tanıtıcı GERÇEKTEN kapanır
                raise OSError(GIZLI)

            wb.close = _close
            return wb

        return mock.patch.object(openpyxl, "load_workbook", _yukle), sayac

    def test_read_xlsx_sheets_okuma_hatasinda_da_kapanir(self):
        cagrilar = self._log_sayaci()
        yol = self._xlsx()
        yama, sayac = self._sayac_close()

        def _patlayan_iter(self_ws, *a, **k):
            raise _hata()

        with yama, mock.patch("openpyxl.worksheet._read_only."
                              "ReadOnlyWorksheet.iter_rows", _patlayan_iter):
            sheets, err = xi._read_xlsx_sheets(str(yol))
        self.assertEqual(len(sayac), 1, "hata yolunda workbook kapatılmadı")
        self.assertEqual(sheets, {})
        self.assertEqual(err, xi.DOSYA_OKUMA_HATASI,
                         "close hatası asıl okuma hatasını maskeledi")
        self.assertEqual(len(cagrilar), 2,
                         f"iki ayrı istisna birer kez loglanmadı: {cagrilar}")
        self._ek_havuz = [err]
        self._sizinti_yok(" (xlsx hata yolu)")

    def test_read_xlsx_sheets_basari_yolunda_tek_kapanis(self):
        yol = self._xlsx()
        yama, sayac = self._sayac_close()
        with yama:
            sheets, err = xi._read_xlsx_sheets(str(yol))
        self.assertEqual(len(sayac), 1, "başarı yolunda tek kapanış değil")
        self.assertEqual(err, "")
        self.assertIn("Müşteriler", sheets)


# ── 16. Kategori yazıldı + ürün fatal: TEK kısmi kutu ───────────────────

class KismiSonucTests(_Temel):

    def _urun_akisi(self, perform):
        yol = self._xlsx()
        return (self._dosya_sec(yol),
                mock.patch.object(xi, "_read_file", staticmethod(
                    lambda *a, **k: ([{"Ürün Kodu": "P1"}], ""))),
                mock.patch.object(xi, "_validate_rows", staticmethod(
                    lambda *a, **k: ([{"product_code": "P1",
                                       "product_name": "Ü1"}], [], []))),
                mock.patch.object(xi, "_perform_import", perform),
                self._onayla(True))

    def _kutu_sayisi(self):
        """Onay kutusu HARİÇ, sonuçta gösterilen kutu sayısı."""
        return len([1 for b, _m in self.kutular if "Onay" not in (b or "")])

    def test_kategori_yazildi_fatal_tek_kutu(self):
        cagrilar = self._log_sayaci()

        def _perform(tur, rows, upd, progress=None, stage_state=None):
            if stage_state is not None:
                stage_state["kategori_yazildi"] = 2
            raise _hata()

        a, b, c, d, e = self._urun_akisi(_perform)
        with a, b, c, d, e:
            sonuc = xi.run_import_flow(None, "products")
        self.assertIs(sonuc, True)
        self.assertEqual(self._kutu_sayisi(), 1,
                         f"tek kısmi kutu değil: {self.kutular}")
        metin = self._metinler()
        self.assertIn("2 kategori oluşturuldu", metin)
        self.assertRegex(metin, r"(?i)ürün kayıtları yazılmadı",
                         "ürünlerin yazılmadığı açıkça söylenmedi")
        self.assertEqual(len(cagrilar), 1, f"tek güvenli log değil: {cagrilar}")
        self._sizinti_yok(" (kısmi kutu)")

    def test_kategori_yokken_normal_hata_ve_false(self):
        def _perform(tur, rows, upd, progress=None, stage_state=None):
            if stage_state is not None:
                stage_state["kategori_yazildi"] = 0
            raise _hata()

        a, b, c, d, e = self._urun_akisi(_perform)
        with a, b, c, d, e:
            sonuc = xi.run_import_flow(None, "products")
        self.assertIs(sonuc, False)
        self.assertEqual(self._kutu_sayisi(), 1)
        self.assertNotIn("kategori oluşturuldu", self._metinler())


# ── 17. Normal sonuçta kategori aşaması görünür ─────────────────────────

class _TekliKarisimi:
    """Tekli ürün akışını sahte `_perform_import` sonucuyla çalıştırır."""

    def _tekli(self, added=0, updated=0, skipped=0, kategori=0, dups=()):
        yol = self._xlsx()

        def _perform(tur, rows, upd, progress=None, stage_state=None):
            if stage_state is not None:
                stage_state["kategori_yazildi"] = kategori
            return added, updated, skipped, []

        with self._dosya_sec(yol), \
             mock.patch.object(xi, "_read_file", staticmethod(
                 lambda *a, **k: ([{"Ürün Kodu": "P1"}], ""))), \
             mock.patch.object(xi, "_validate_rows", staticmethod(
                 lambda *a, **k: ([{"product_code": "P1",
                                    "product_name": "Ü"}], list(dups), []))), \
             mock.patch.object(xi, "_perform_import", _perform), \
             self._onayla(True):
            return xi.run_import_flow(None, "products")


class KategoriGorunurlukTests(_Temel, _TekliKarisimi):

    def test_normal_donuste_kategori_gorunur(self):
        sonuc = self._tekli(added=0, updated=0, kategori=1)
        metin = self._metinler()
        self.assertIn("1 kategori oluşturuldu", metin)
        self.assertNotIn("İşlem yapılmadı", metin,
                         "kategori yazıldığı hâlde 'işlem yapılmadı' dendi")
        self.assertIs(sonuc, True)

    def test_tumunu_aktar_normal_yolda_kategori_gorunur(self):
        yol = self._xlsx()

        def _perform(tur, rows, upd, progress=None, stage_state=None):
            if stage_state is not None:
                stage_state["kategori_yazildi"] = 2
            return 0, 0, 0, []

        with self._dosya_sec(yol), \
             mock.patch.object(xi, "_read_xlsx_sheets", staticmethod(
                 lambda *a, **k: ({"Ürünler": [{"Ürün Kodu": "P1"}]}, ""))), \
             mock.patch.object(xi, "_validate_rows", staticmethod(
                 lambda *a, **k: ([{"product_code": "P1",
                                    "product_name": "Ü"}], [], []))), \
             mock.patch.object(xi, "_validate_offer_rows", staticmethod(
                 lambda *a, **k: ([], [], []))), \
             mock.patch.object(xi, "_perform_import", _perform), \
             self._onayla(True):
            sonuc = xi.run_import_all_flow(None)
        self.assertRegex(self._metinler(), r"Kategori:\s*2 oluşturuldu")
        self.assertIs(sonuc, True)

    def test_tumunu_aktar_fatal_yolda_kategori_ve_hata_birlikte(self):
        yol = self._xlsx()

        def _perform(tur, rows, upd, progress=None, stage_state=None):
            if stage_state is not None:
                stage_state["kategori_yazildi"] = 2
            raise _hata()

        with self._dosya_sec(yol), \
             mock.patch.object(xi, "_read_xlsx_sheets", staticmethod(
                 lambda *a, **k: ({"Ürünler": [{"Ürün Kodu": "P1"}]}, ""))), \
             mock.patch.object(xi, "_validate_rows", staticmethod(
                 lambda *a, **k: ([{"product_code": "P1",
                                    "product_name": "Ü"}], [], []))), \
             mock.patch.object(xi, "_validate_offer_rows", staticmethod(
                 lambda *a, **k: ([], [], []))), \
             mock.patch.object(xi, "_perform_import", _perform), \
             self._onayla(True):
            sonuc = xi.run_import_all_flow(None)
        metin = self._metinler()
        self.assertRegex(metin, r"Kategori:\s*2 oluşturuldu")
        self.assertRegex(metin, r"(?i)ürün.*aktarılamadı")
        self.assertIs(sonuc, True)
        self._sizinti_yok(" (tümünü kategori + fatal)")


# ── 18. Tekli import dönüş değeri = gerçek DB değişikliği ───────────────

class DonusDegeriTests(_Temel, _TekliKarisimi):

    def test_hicbir_degisiklik_yoksa_false(self):
        self.assertIs(self._tekli(added=0, updated=0, skipped=3, kategori=0),
                      False, "DB değişmediği hâlde cache yenilemesi istendi")

    def test_yalniz_mukerrer_atlandiysa_false(self):
        self.assertIs(self._tekli(added=0, updated=0, kategori=0,
                                  dups=[{"product_code": "D1"}]),
                      False)

    def test_ekleme_varsa_true(self):
        self.assertIs(self._tekli(added=1), True)

    def test_guncelleme_varsa_true(self):
        self.assertIs(self._tekli(updated=1), True)

    def test_yalniz_kategori_yazildiysa_true(self):
        self.assertIs(self._tekli(added=0, updated=0, kategori=1), True)


# ── 9. Kaynak koruması ──────────────────────────────────────────────────

class KaynakKorumasiTests(unittest.TestCase):

    def test_ham_istisna_gosterimi_yok(self):
        kaynak = inspect.getsource(xi)
        for yasak in ("{e}", "{exc}", "exc_info=True", "str(e)", "str(exc)"):
            self.assertNotIn(yasak, kaynak, f"ham istisna gösterimi: {yasak}")

    def test_errors_listesine_kullanici_verisi_eklenmiyor(self):
        for fn in (xi._perform_import, xi._perform_offer_import):
            for satir in inspect.getsource(fn).splitlines():
                if "errors.append" in satir:
                    for yasak in ("{e}", "{exc}", "company", "code", "offer_no"):
                        self.assertNotIn(yasak, satir,
                                         f"errors'a kullanıcı verisi: {satir}")


# ── test altyapısı: gerçek transaction'ı saran yardımcı ─────────────────

def xi_db_transaction():
    from database.db_manager import get_db
    return get_db().transaction


class _sarmalayan_transaction:
    """Gerçek transaction'ı kullanır ama seçilen `execute` çağrılarını patlatır.

    Böylece rollback/commit davranışı GERÇEK kalır; yalnız tek satırın
    yazımı başarısız olur.
    """

    def __init__(self, patlayanlar):
        self.patlayanlar = patlayanlar
        self._patch = None

    def __enter__(self):
        from database.db_manager import get_db
        db = get_db()
        gercek = db.transaction
        durum = {"n": 0}
        patlayanlar = self.patlayanlar

        class _Conn:
            def __init__(self, ic):
                self._ic = ic

            def execute(self, *a, **k):
                i = durum["n"]
                durum["n"] += 1
                if i in patlayanlar:
                    raise RuntimeError(GIZLI)
                return self._ic.execute(*a, **k)

            def __getattr__(self, ad):
                return getattr(self._ic, ad)

        import contextlib

        @contextlib.contextmanager
        def _tx(*a, **k):
            with gercek(*a, **k) as conn:
                yield _Conn(conn)

        self._patch = mock.patch.object(type(db), "transaction", _tx)
        self._patch.start()
        return self

    def __exit__(self, *a):
        self._patch.stop()
        return False


if __name__ == "__main__":
    unittest.main(verbosity=2)
