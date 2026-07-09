"""ExportService birim testleri."""
import csv
import tempfile
import unittest
from datetime import date
from pathlib import Path

from database.db_manager import get_db
from models.offer import Offer
from models.offer_item import OfferItem
from services.export_service import export_csv, export_excel
from services.offer_service import OfferService


def _offer(currency="TL", amount=90.0):
    return Offer(
        company_name="Export Firma",
        date=date.today().strftime("%d.%m.%Y"),
        currency=currency,
        total_amount=amount,
        discount_amount=10.0,
        discount_type="percent",
        discount_value=10.0,
        items=[OfferItem(
            product_name="Export Ürün",
            quantity=1,
            unit_price=100.0,
            total_price=100.0,
        )],
    )


class TestExportService(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.db = get_db()
        cls.svc = OfferService()

    @classmethod
    def tearDownClass(cls):
        pass

    def setUp(self):
        with self.db.transaction() as conn:
            conn.execute("DELETE FROM offer_items")
            conn.execute("DELETE FROM offers")
            conn.execute("DELETE FROM offer_counter")

    def _save_and_get(self, currency="TL", amount=90.0):
        oid = self.svc.save(_offer(currency, amount))
        return self.svc.get_by_id(oid)

    # ── Excel ────────────────────────────────────────────────────────────

    def test_excel_creates_file(self):
        offers = [self._save_and_get()]
        with tempfile.TemporaryDirectory() as td:
            path = str(Path(td) / "test.xlsx")
            result = export_excel(offers, path)
            self.assertTrue(Path(result).exists())
            self.assertGreater(Path(result).stat().st_size, 0)

    def test_excel_has_correct_headers(self):
        offers = [self._save_and_get()]
        with tempfile.TemporaryDirectory() as td:
            path = str(Path(td) / "headers.xlsx")
            export_excel(offers, path)
            from openpyxl import load_workbook
            ws = load_workbook(path).active
            headers = [ws.cell(1, c).value for c in range(1, 10)]
            self.assertIn("Teklif No", headers)
            self.assertIn("Firma", headers)
            self.assertIn("Toplam Tutar", headers)

    def test_excel_multi_currency_totals(self):
        offers = [self._save_and_get("TL"), self._save_and_get("EUR")]
        with tempfile.TemporaryDirectory() as td:
            path = str(Path(td) / "multi.xlsx")
            export_excel(offers, path)
            from openpyxl import load_workbook
            ws = load_workbook(path).active
            summary_row = len(offers) + 2
            total_cell = ws.cell(summary_row, 5).value
            self.assertIn("|", total_cell)

    # ── CSV ──────────────────────────────────────────────────────────────

    def test_csv_creates_file(self):
        offers = [self._save_and_get()]
        with tempfile.TemporaryDirectory() as td:
            path = str(Path(td) / "test.csv")
            result = export_csv(offers, path)
            self.assertTrue(Path(result).exists())

    def test_csv_utf8_sig_encoding(self):
        offers = [self._save_and_get()]
        with tempfile.TemporaryDirectory() as td:
            path = str(Path(td) / "encoding.csv")
            export_csv(offers, path)
            raw = Path(path).read_bytes()
            self.assertTrue(raw.startswith(b"\xef\xbb\xbf"))

    def test_csv_semicolon_delimiter(self):
        offers = [self._save_and_get()]
        with tempfile.TemporaryDirectory() as td:
            path = str(Path(td) / "delim.csv")
            export_csv(offers, path)
            with open(path, encoding="utf-8-sig") as f:
                reader = csv.reader(f, delimiter=";")
                rows = list(reader)
            self.assertEqual(len(rows), 2)
            self.assertGreater(len(rows[0]), 1)

    def test_csv_correct_row_count(self):
        offers = [self._save_and_get(), self._save_and_get("EUR")]
        with tempfile.TemporaryDirectory() as td:
            path = str(Path(td) / "count.csv")
            export_csv(offers, path)
            with open(path, encoding="utf-8-sig") as f:
                rows = list(csv.reader(f, delimiter=";"))
            self.assertEqual(len(rows), 3)

    def test_export_empty_list(self):
        with tempfile.TemporaryDirectory() as td:
            xlsx_path = str(Path(td) / "empty.xlsx")
            csv_path = str(Path(td) / "empty.csv")
            export_excel([], xlsx_path)
            export_csv([], csv_path)
            self.assertTrue(Path(xlsx_path).exists())
            self.assertTrue(Path(csv_path).exists())


if __name__ == "__main__":
    unittest.main(verbosity=2)
