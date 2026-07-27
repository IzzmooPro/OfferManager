"""O15 — çok sayfalı XLSX'te sayfa seçimi (D seçeneği).

Ölçüm: `_read_file` yalnız `workbook.active` sayfasını okuyordu; diğer
sayfalardaki geçerli veri kullanıcıya bildirilmeden atılıyordu. Gerçek
kullanıcı dosyalarından `Urunler_2_Cetinkaya_Pano_HAZIR.xlsx` iki uygun
sayfa içeriyor.

Sözleşme: tek geçerli GÖRÜNÜR veri sayfası otomatik seçilir (ek tıklama yok);
birden fazlaysa kullanıcıya sorulur; gizli sayfalar ne listelenir ne okunur;
yalnız seçilen sayfa aktarılır (otomatik birleştirme YOK).
"""
import csv
import os
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import mock

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import Workbook
from PySide6.QtWidgets import QApplication

from database.db_manager import get_db
from ui.utils import excel_import as ei

U_BAS = ["Ürün Kodu", "Ürün Adı", "Fiyat"]
M_BAS = ["Firma Adı", "Telefon"]


class _Temel(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.app = QApplication.instance() or QApplication([])
        cls.db = get_db()

    def setUp(self):
        self._tmp = TemporaryDirectory(prefix="o15t_", ignore_cleanup_errors=True)
        self.addCleanup(self._tmp.cleanup)
        self.kok = Path(self._tmp.name)
        self._sayac = 0
        with self.db.transaction() as conn:
            conn.execute("DELETE FROM products")
            conn.execute("DELETE FROM customers")

    def _xlsx(self, sayfalar, aktif=None, gizli=()):
        self._sayac += 1
        wb = Workbook()
        wb.remove(wb.active)
        for ad, satirlar in sayfalar:
            ws = wb.create_sheet(ad)
            for s in satirlar:
                ws.append(s)
            if ad in gizli:
                ws.sheet_state = "hidden"
        if aktif:
            wb.active = wb.sheetnames.index(aktif)
        yol = self.kok / f"d{self._sayac}.xlsx"
        wb.save(yol)
        return yol

    def _oku(self, yol, tur="products", secim=None):
        """(_read_file sonucu, getItem çağrı sayısı, gösterilen adaylar)"""
        cagri = {"n": 0, "items": None, "current": None, "mesaj": None}

        def sahte_getItem(parent, baslik, mesaj, items, current=0,
                          editable=False, *a, **k):
            cagri["n"] += 1
            cagri["items"] = list(items)
            cagri["current"] = current
            cagri["mesaj"] = mesaj
            if secim is None:
                return "", False                    # İptal
            return items[secim], True

        with mock.patch("PySide6.QtWidgets.QInputDialog.getItem",
                        side_effect=sahte_getItem):
            satirlar, hata = ei._read_file(str(yol), import_type=tur,
                                           parent=None)
        return (satirlar, hata), cagri


class SingleSheetTests(_Temel):

    def test_single_product_sheet_no_dialog(self):
        yol = self._xlsx([("Ürünler", [U_BAS, ["A-1", "Bir", 1],
                                       ["A-2", "İki", 2]])])
        (satir, hata), c = self._oku(yol)
        self.assertEqual(hata, "")
        self.assertEqual(len(satir), 2)
        self.assertEqual(c["n"], 0, "tek sayfada dialog açıldı")

    def test_single_customer_sheet_no_dialog(self):
        yol = self._xlsx([("Müşteriler", [M_BAS, ["Acme", "1"]])])
        (satir, hata), c = self._oku(yol, tur="customers")
        self.assertEqual(len(satir), 1)
        self.assertEqual(c["n"], 0)


class AutoSelectTests(_Temel):

    def test_empty_active_sheet_falls_to_valid_one(self):
        yol = self._xlsx([("Boş", [U_BAS]),
                          ("Veri", [U_BAS, ["V-1", "Var", 1],
                                    ["V-2", "Var2", 2]])])
        (satir, hata), c = self._oku(yol)
        self.assertEqual(c["n"], 0, "tek aday varken dialog açıldı")
        self.assertEqual(len(satir), 2, "aktif olmayan geçerli sayfa okunmadı")

    def test_instruction_active_sheet_falls_to_valid_one(self):
        yol = self._xlsx([("Açıklama", [["Bu dosyayı doldurun"], ["not"]]),
                          ("Ürünler", [U_BAS, ["T-1", "Gerçek", 5]])])
        (satir, hata), c = self._oku(yol)
        self.assertEqual(c["n"], 0)
        self.assertEqual(len(satir), 1)
        self.assertEqual(satir[0]["Ürün Kodu"], "T-1")

    def test_customer_import_picks_customer_sheet(self):
        yol = self._xlsx([("Ürünler", [U_BAS, ["P-1", "Ürün", 1]]),
                          ("Müşteriler", [M_BAS, ["Acme", "1"]])])
        (satir, hata), c = self._oku(yol, tur="customers")
        self.assertEqual(c["n"], 0)
        self.assertEqual(len(satir), 1)
        self.assertIn("Firma Adı", satir[0])

    def test_product_import_picks_product_sheet(self):
        yol = self._xlsx([("Müşteriler", [M_BAS, ["Acme", "1"]]),
                          ("Ürünler", [U_BAS, ["P-1", "Ürün", 1]])])
        (satir, hata), c = self._oku(yol, tur="products")
        self.assertEqual(c["n"], 0)
        self.assertIn("Ürün Kodu", satir[0])


class MultiCandidateTests(_Temel):

    def _iki_sayfa(self, aktif=None, gizli=()):
        return self._xlsx([("S1", [U_BAS, ["S1-1", "Bir", 1]]),
                           ("S2", [U_BAS, ["S2-1", "İki", 2],
                                   ["S2-2", "Üç", 3]])],
                          aktif=aktif, gizli=gizli)

    def test_dialog_shown_exactly_once(self):
        (satir, hata), c = self._oku(self._iki_sayfa(), secim=0)
        self.assertEqual(c["n"], 1, "seçim dialogu tam bir kez açılmalı")
        self.assertEqual(len(c["items"]), 2)

    def test_selected_second_sheet_is_read(self):
        (satir, hata), c = self._oku(self._iki_sayfa(), secim=1)
        self.assertEqual(len(satir), 2, "seçilen sayfa okunmadı")
        self.assertEqual([r["Ürün Kodu"] for r in satir], ["S2-1", "S2-2"])

    def test_other_sheets_are_not_merged(self):
        (satir, hata), c = self._oku(self._iki_sayfa(), secim=0)
        self.assertEqual(len(satir), 1, "sayfalar otomatik birleştirildi")
        self.assertEqual(satir[0]["Ürün Kodu"], "S1-1")

    def test_default_selection_is_active_sheet(self):
        (satir, hata), c = self._oku(self._iki_sayfa(aktif="S2"), secim=0)
        self.assertTrue(c["items"][c["current"]].startswith("S2"),
                        "varsayılan seçim aktif sayfa değil")

    def test_dialog_text_explains_single_sheet_import(self):
        (satir, hata), c = self._oku(self._iki_sayfa(), secim=0)
        self.assertIn("Yalnız seçtiğiniz sayfa", c["mesaj"])

    def test_dialog_items_show_approximate_row_counts(self):
        """Sayı YAKLAŞIKTIR; kesin sayım için ikinci tarama yapılmaz."""
        (satir, hata), c = self._oku(self._iki_sayfa(), secim=0)
        for metin in c["items"]:
            self.assertIn("yaklaşık", metin.lower())
            self.assertIn("satır", metin.lower())

    def test_unknown_dialog_result_is_safe_cancel(self):
        """QInputDialog beklenmedik metin dönerse ham ValueError sızmamalı."""
        yol = self._iki_sayfa()
        with mock.patch("PySide6.QtWidgets.QInputDialog.getItem",
                        return_value=("BEKLENMEDIK", True)):
            satir, hata = ei._read_file(str(yol), import_type="products")
        self.assertEqual(satir, [])
        self.assertEqual(hata, ei.SAYFA_SECIMI_IPTAL)

    def test_three_candidates_listed(self):
        yol = self._xlsx([("A", [U_BAS, ["A-1", "a", 1]]),
                          ("B", [U_BAS, ["B-1", "b", 1]]),
                          ("C", [U_BAS, ["C-1", "c", 1]])])
        (satir, hata), c = self._oku(yol, secim=2)
        self.assertEqual(len(c["items"]), 3)
        self.assertEqual(satir[0]["Ürün Kodu"], "C-1")

    def test_unicode_sheet_names_shown(self):
        yol = self._xlsx([("Ürün Listesi", [U_BAS, ["Ü-1", "Ürün", 1]]),
                          ("Şablon", [U_BAS, ["Ş-1", "Şablon", 2]])])
        (satir, hata), c = self._oku(yol, secim=1)
        self.assertTrue(any("Şablon" in m for m in c["items"]))
        self.assertEqual(satir[0]["Ürün Kodu"], "Ş-1")


class CancelTests(_Temel):

    def test_cancel_returns_no_rows_and_no_error_dialog(self):
        yol = self._xlsx([("S1", [U_BAS, ["S1-1", "a", 1]]),
                          ("S2", [U_BAS, ["S2-1", "b", 2]])])
        (satir, hata), c = self._oku(yol, secim=None)
        self.assertEqual(satir, [])
        self.assertEqual(hata, ei.SAYFA_SECIMI_IPTAL)

    def test_cancel_aborts_flow_without_db_write_or_message(self):
        yol = self._xlsx([("S1", [U_BAS, ["S1-1", "a", 1]]),
                          ("S2", [U_BAS, ["S2-1", "b", 2]])])
        with mock.patch("PySide6.QtWidgets.QFileDialog.getOpenFileName",
                        return_value=(str(yol), "")), \
             mock.patch("PySide6.QtWidgets.QInputDialog.getItem",
                        return_value=("", False)), \
             mock.patch.object(ei, "_validate_rows") as dogrula, \
             mock.patch.object(ei, "_perform_import") as yaz, \
             mock.patch("ui.utils.excel_import.QMessageBox") as kutu:
            sonuc = ei.run_import_flow(None, "products")
        self.assertFalse(sonuc)
        dogrula.assert_not_called()
        yaz.assert_not_called()
        kutu.warning.assert_not_called()
        kutu.information.assert_not_called()
        self.assertEqual(
            self.db.fetchone("SELECT COUNT(*) c FROM products")["c"], 0)


class HiddenSheetTests(_Temel):

    def test_hidden_sheet_not_listed_and_not_read(self):
        yol = self._xlsx([("Görünür", [U_BAS, ["G-1", "a", 1]]),
                          ("Gizli", [U_BAS, ["Z-1", "b", 2]])],
                         gizli=("Gizli",))
        (satir, hata), c = self._oku(yol)
        self.assertEqual(c["n"], 0, "gizli sayfa aday sayıldı")
        self.assertEqual([r["Ürün Kodu"] for r in satir], ["G-1"])

    def test_only_hidden_valid_sheet_gives_explicit_error(self):
        yol = self._xlsx([("Boş", [U_BAS]),
                          ("Gizli", [U_BAS, ["Z-1", "b", 2]])],
                         gizli=("Gizli",))
        (satir, hata), c = self._oku(yol)
        self.assertEqual(satir, [])
        self.assertNotEqual(hata, "")
        self.assertIn("gizli", hata.lower())
        self.assertNotIn("Traceback", hata)

    def test_hidden_notice_in_dialog_when_candidates_exist(self):
        yol = self._xlsx([("S1", [U_BAS, ["S1-1", "a", 1]]),
                          ("S2", [U_BAS, ["S2-1", "b", 2]]),
                          ("Gizli", [U_BAS, ["Z-1", "c", 3]])],
                         gizli=("Gizli",))
        (satir, hata), c = self._oku(yol, secim=0)
        self.assertEqual(len(c["items"]), 2)
        self.assertIn("Gizli çalışma sayfaları içe aktarılmaz", c["mesaj"])


class NoCandidateTests(_Temel):

    def test_header_only_and_empty_sheets_are_not_candidates(self):
        yol = self._xlsx([("Başlık", [U_BAS]), ("Boş", []),
                          ("Not", [["açıklama"], ["ikinci"]])])
        (satir, hata), c = self._oku(yol)
        self.assertEqual(satir, [])
        self.assertNotEqual(hata, "")
        self.assertNotIn("Traceback", hata)

    def test_wrong_type_sheet_is_not_candidate(self):
        yol = self._xlsx([("Müşteriler", [M_BAS, ["Acme", "1"]])])
        (satir, hata), c = self._oku(yol, tur="products")
        self.assertEqual(satir, [])
        self.assertNotEqual(hata, "")


class SourceSheetMetadataTests(_Temel):

    def test_rows_carry_source_sheet(self):
        yol = self._xlsx([("Ürün Listesi", [U_BAS, ["A-1", "a", 1]])])
        (satir, hata), c = self._oku(yol)
        self.assertEqual(satir[0].get("_source_sheet"), "Ürün Listesi")

    def test_validator_error_mentions_sheet_name(self):
        yol = self._xlsx([("Ürün Listesi", [U_BAS, ["A-1", "a", 1],
                                            ["", "Kodsuz", 2]])])
        (satir, hata), c = self._oku(yol)
        valid, dup, invalid = ei._validate_rows("products", satir)
        self.assertEqual(len(invalid), 1)
        self.assertIn("Ürün Listesi", invalid[0]["_error"])
        self.assertNotIn("_source_sheet", invalid[0]["_error"])

    def test_csv_rows_have_no_sheet_suffix(self):
        yol = self.kok / "u.csv"
        with open(yol, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(U_BAS)
            w.writerow(["", "Kodsuz", 1])
        satir, hata = ei._read_file(str(yol), import_type="products")
        self.assertEqual(hata, "")
        valid, dup, invalid = ei._validate_rows("products", satir)
        self.assertEqual(len(invalid), 1)
        self.assertNotIn("sayfa", invalid[0]["_error"].lower())


class WorkbookCloseTests(_Temel):
    """Workbook TÜM yollarda kapatılmalı."""

    def _izle(self, yol, tur="products", getitem=None, bozuk=None):
        """(kapanis_sayisi, sonuc) — load_workbook sarmalanır."""
        import openpyxl
        orj = openpyxl.load_workbook
        kayit = {"acilan": 0, "kapanan": 0}

        class _Sarmal:
            def __init__(self, wb):
                self._wb = wb

            def close(self):
                kayit["kapanan"] += 1
                return self._wb.close()

            def __getattr__(self, ad):
                return getattr(self._wb, ad)

            def __getitem__(self, ad):
                return self._wb[ad]

        def sahte_load(*a, **k):
            kayit["acilan"] += 1
            return _Sarmal(orj(*a, **k))

        yamalar = [mock.patch.object(openpyxl, "load_workbook", sahte_load),
                   mock.patch.object(ei, "_sayfa_adaylari",
                                     side_effect=bozuk) if bozuk else None]
        if getitem is not None:
            yamalar.append(mock.patch(
                "PySide6.QtWidgets.QInputDialog.getItem", return_value=getitem))
        yamalar = [y for y in yamalar if y is not None]
        for y in yamalar:
            y.start()
        try:
            try:
                sonuc = ei._read_file(str(yol), import_type=tur)
            except Exception as e:
                sonuc = ("ISTISNA", type(e).__name__)
        finally:
            for y in yamalar:
                y.stop()
        return kayit, sonuc

    def _tek(self):
        return self._xlsx([("Ürünler", [U_BAS, ["A-1", "a", 1]])])

    def _cok(self):
        return self._xlsx([("S1", [U_BAS, ["S1-1", "a", 1]]),
                           ("S2", [U_BAS, ["S2-1", "b", 2]])])

    def test_closed_on_single_candidate_success(self):
        k, _ = self._izle(self._tek())
        self.assertEqual((k["acilan"], k["kapanan"]), (1, 1))

    def test_closed_on_multi_candidate_selection(self):
        k, _ = self._izle(self._cok(),
                          getitem=("S2 — yaklaşık 1 satır", True))
        self.assertEqual((k["acilan"], k["kapanan"]), (1, 1))

    def test_closed_on_user_cancel(self):
        k, sonuc = self._izle(self._cok(), getitem=("", False))
        self.assertEqual((k["acilan"], k["kapanan"]), (1, 1))
        self.assertEqual(sonuc[1], ei.SAYFA_SECIMI_IPTAL)

    def test_closed_when_no_candidate(self):
        yol = self._xlsx([("Başlık", [U_BAS])])
        k, _ = self._izle(yol)
        self.assertEqual((k["acilan"], k["kapanan"]), (1, 1))

    def test_closed_when_candidate_scan_raises(self):
        k, sonuc = self._izle(self._tek(),
                              bozuk=RuntimeError("tarama patladı"))
        self.assertEqual((k["acilan"], k["kapanan"]), (1, 1))
        self.assertNotEqual(sonuc[1], ei.SAYFA_SECIMI_IPTAL)

    def test_closed_when_progress_callback_raises(self):
        yol = self._tek()
        import openpyxl
        orj = openpyxl.load_workbook
        kayit = {"kapanan": 0}

        class _S:
            def __init__(self, wb):
                self._wb = wb

            def close(self):
                kayit["kapanan"] += 1
                return self._wb.close()

            def __getattr__(self, ad):
                return getattr(self._wb, ad)

            def __getitem__(self, ad):
                return self._wb[ad]

        def patlayan_progress(*a, **k):
            raise RuntimeError("progress patladı")

        with mock.patch.object(openpyxl, "load_workbook",
                               lambda *a, **k: _S(orj(*a, **k))):
            satir, hata = ei._read_file(str(yol), progress=patlayan_progress,
                                        import_type="products")
        self.assertEqual(kayit["kapanan"], 1, "progress hatasında kapanmadı")
        self.assertNotEqual(hata, ei.SAYFA_SECIMI_IPTAL)


class CancelSentinelScopeTests(_Temel):
    """Sentinel YALNIZ kullanıcı iptalinde üretilmeli."""

    def test_read_error_is_not_reported_as_cancel(self):
        bozuk = self.kok / "bozuk.xlsx"
        bozuk.write_bytes(b"bu bir xlsx degil")
        satir, hata = ei._read_file(str(bozuk), import_type="products")
        self.assertEqual(satir, [])
        self.assertNotEqual(hata, ei.SAYFA_SECIMI_IPTAL)
        self.assertNotEqual(hata, "")

    def test_no_candidate_is_not_reported_as_cancel(self):
        yol = self._xlsx([("Başlık", [U_BAS])])
        satir, hata = ei._read_file(str(yol), import_type="products")
        self.assertNotEqual(hata, ei.SAYFA_SECIMI_IPTAL)


class ScanEfficiencyTests(_Temel):
    """Aday taraması ilk gerçek veri satırından sonra durmalı."""

    def test_scan_stops_after_first_data_row(self):
        satirlar = [U_BAS] + [[f"K-{i}", f"Ad {i}", i] for i in range(300)]
        yol = self._xlsx([("Büyük", satirlar)])
        import openpyxl
        wb = openpyxl.load_workbook(yol, read_only=True, data_only=True)
        okunan = {"n": 0}
        try:
            ws = wb.worksheets[0]
            orj_iter = ws.iter_rows          # BAĞLI metot

            def sayan(*a, **k):
                for r in orj_iter(*a, **k):
                    okunan["n"] += 1
                    yield r

            ws.iter_rows = sayan             # örnek düzeyinde yama
            adaylar, gizli = ei._sayfa_adaylari(wb, "products")
        finally:
            wb.close()
        self.assertEqual(len(adaylar), 1)
        self.assertLessEqual(okunan["n"], 5,
                             f"aday taraması {okunan['n']} satır okudu")


class MetadataLeakTests(_Temel):
    """`_source_sheet` DB alanlarına sızmamalı."""

    def test_source_sheet_not_written_to_products(self):
        yol = self._xlsx([("Ürünler", [U_BAS, ["MD-1", "Ürün", 5]])])
        satir, hata = ei._read_file(str(yol), import_type="products")
        valid, dup, invalid = ei._validate_rows("products", satir)
        ei._perform_import("products", list(valid), False)
        kolonlar = [r["name"] for r in
                    self.db.fetchall("PRAGMA table_info(products)")]
        self.assertNotIn("_source_sheet", kolonlar)
        kayit = self.db.fetchone(
            "SELECT * FROM products WHERE product_code='MD-1'")
        self.assertIsNotNone(kayit)
        self.assertNotIn("_source_sheet", dict(kayit))

    def test_source_sheet_not_written_to_customers(self):
        yol = self._xlsx([("Müşteriler", [M_BAS, ["Meta AS", "1"]])])
        satir, hata = ei._read_file(str(yol), import_type="customers")
        valid, dup, invalid = ei._validate_rows("customers", satir)
        ei._perform_import("customers", list(valid), False)
        kayit = self.db.fetchone(
            "SELECT * FROM customers WHERE company_name='Meta AS'")
        self.assertIsNotNone(kayit)
        self.assertNotIn("_source_sheet", dict(kayit))


class HiddenNoticeWithSingleCandidateTests(_Temel):

    def test_notice_shown_when_single_visible_candidate(self):
        """Tek görünür aday + gizli uygun sayfa → not kullanıcıya ulaşmalı."""
        yol = self._xlsx([("Görünür", [U_BAS, ["G-1", "a", 1]]),
                          ("Gizli", [U_BAS, ["Z-1", "b", 2]])],
                         gizli=("Gizli",))
        import openpyxl
        wb = openpyxl.load_workbook(yol, read_only=True, data_only=True)
        try:
            adaylar, gizli_dolu = ei._sayfa_adaylari(wb, "products")
        finally:
            wb.close()
        self.assertEqual(len(adaylar), 1)
        self.assertTrue(gizli_dolu, "gizli uygun sayfa bayrağı kurulmadı")
        self.assertIn("Gizli çalışma sayfaları içe aktarılmaz",
                      ei._gizli_notu(gizli_dolu))


class ImportAllUnchangedTests(_Temel):
    """`_read_xlsx_sheets` (Tümünü İçe Aktar) bu turda değişmemeli."""

    def test_reads_all_sheets_including_hidden(self):
        yol = self._xlsx([("Müşteriler", [M_BAS, ["Acme", "1"]]),
                          ("Ürünler", [U_BAS, ["P-1", "Ürün", 1]]),
                          ("Gizli", [U_BAS, ["Z-1", "b", 2]])],
                         gizli=("Gizli",))
        sheets, err = ei._read_xlsx_sheets(str(yol))
        self.assertEqual(err, "")
        self.assertEqual(set(sheets), {"Müşteriler", "Ürünler", "Gizli"})


if __name__ == "__main__":
    unittest.main()
